# ===================================================================================================
# Script para extração e exportação dos dados contidos nas Notas de Corretagem no padrão SINACOR
# Os dados extraídos são inseridos em planilhas excel
# O atual scritp foi testado apenas nas corretoras XP, Rico e Agora
# Para outras corretoras favor enviar notas de corretagem para implementação
# Para dúvidas e sugestões entrar em contato pelo e-mail: marcelo.pcf@gmail.com
# ===================================================================================================

from os.path import isfile, join, basename, exists
import sys
import re
import shutil
from datetime import datetime
#from os import listdir, makedirs
#import subprocess
#import os
#from shutil import copytree, ignore_patterns
#from scipy.special import logsumexp
#import time

# ====================================================================================================
# Verifica se está rodando versão correta do Python
# ====================================================================================================
if sys.version_info <= (3, 0):
    sys.stdout.write("Versao do intepretador python (" + str(platform.python_version()) + ") inadequada.\n")
    sys.stdout.write("Este programa requer Python 3 (preferencialmente Python 3.9.2 ou superior).\n")
    sys.exit(1)

# ===================================================================================================
# Carga de módulos opcionais
# ===================================================================================================
def instalar_modulo(modulo):
    # Comando para instalar módulos
    import subprocess
    comando = sys.executable + " -m" + " pip" + " install " + modulo
    print("-"*100)
    print("- O módulo", modulo,"não vem embutido na instalação do python e necessita de instalação específica.")
    print("- Instalando módulo opcional: ", modulo, "Aguarde....")
    subprocess.call([sys.executable, "-m", "pip", "install", modulo])
    if modulo == 'tabula-py':
        modulo = 'tabula'
    try:
        __import__(modulo)
    except ImportError as e:
        print("- Erro: Instalação de Módulo adicional", modulo, "falhou: " + str(e))
        print("- Para efetar uma instalação manual, conecte este computador na internet e utilize o comando abaixo")
        print(comando)
        input("- Digite <ENTER> para prosseguir")
        sys.exit(1)
    return comando

# ===================================================================================================
# Lista de modulos opcionais
# ATENÇÃO: PARA CADA MÓDULO NOVO, INCLUIR AS DUAS LINHAS, com a definição da váriavel modulo e o import
# ===================================================================================================
modulo = ''
try:
    modulo='shutil'
    import  shutil
except ImportError as e:
    print("-"*100)
    print(str(e))
    comando=instalar_modulo(modulo)
    
try:
    modulo='pandas'
    import  pandas
except ImportError as e:
    print("-"*100)
    print(str(e))
    comando=instalar_modulo(modulo)
    
try:
    modulo='openpyxl'
    import  openpyxl
except ImportError as e:
    print("-"*100)
    print(str(e))
    comando=instalar_modulo(modulo)
    
try:
    modulo='xlwings'
    import  xlwings
except ImportError as e:
    print("-"*100)
    print(str(e))
    comando=instalar_modulo(modulo)
    
try:
    modulo='tabula-py'
    import tabula
except ImportError as e:
    print("-"*100)
    print(str(e))
    comando=instalar_modulo(modulo)

### Por enquanto não são necessários
#try:
#    modulo='et-xmlfile'
#    import  et-xmlfile
#except ImportError as e:
#    print("-"*100)
#    print(str(e))
#    comando=instalar_modulo(modulo)
#
# Talvez precise mais tarde
# try:
#     modulo='cryptography'
#     import cryptography
# except ImportError as e:
#     print("-"*100)
#     print(str(e))
#     comando=instalar_modulo(modulo)
#
# try:
#     modulo='python-dateutil'
#     import python-dateutil
# except ImportError as e:
#     print("-"*100)
#     print(str(e))
#     comando=instalar_modulo(modulo)
# try:
#     modulo='scipy'
#     import  scipy
# except ImportError as e:
#     print("-"*100)
#     print(str(e))
#     comando=instalar_modulo(modulo)
#
#
# try:
#     modulo='numpy'
#     import  numpy
# except ImportError as e:
#     print("-"*100)
#     print(str(e))
#     comando=instalar_modulo(modulo)
#
# try:
#     modulo='sklearn'
#     import  sklearn
# except ImportError as e:
#     print("-"*100)
#     print(str(e))
#     comando=instalar_modulo(modulo)

#Simula um erro de instalção de módulo
#try:
#    modulo='xxxxxxx'
#    import  xxxxxxx
#except ImportError as e:
#    print("-"*100)
#    print(str(e))
#    comando=instalar_modulo(modulo)

# ===================================================================================================
# Modulos necessários para execução correta do script de extração de dados das Notas de Corretagem
# ATENÇÃO: Esses modulos serão instalados automaticamente, 
#          caso não estejam instalados na primeira execução do script.
# ===================================================================================================
from tabula import read_pdf 
import pandas as pd
#import xlwings as xw
#from openpyxl import load_workbook
#from openpyxl.worksheet.table import Table, TableStyleInfo

# ===================================================================================================
# "Desenhos" para direcionar a atenção ou erro
# ===================================================================================================
def print_atencao():
    # Na tela sai legal...aqui está distorcido, provavelmente em função da largura dos caracteres
    # teria que ter uma fonte com largura fixa
    print("┌─────────────────┐")
    print("│  A T E N Ç Ã O  │")
    print("└─────────────────┘")

def print_erro():
    # Na tela sai legal...aqui está distorcido, provavelmente em função da largura dos caracteres
    # teria que ter uma fonte com largura fixa
    print("┌───────────┐")
    print("│  E R R O  │")
    print("└───────────┘")

# unicode box drawing characteres
'''
┌ ┐
└ ┘
─
│
┴
├ ┤
┬
╷
'┼'
'''
# ===================================================================================================
# Arquivo CSV contendo o nome dos papeis no pregão da B3
# ATENÇÃO: Para cada NOVO papel, incluir o nome correto no PREGÃO e o CÓDIGO correspondente na B3
# ===================================================================================================

# acoes = pd.read_csv('./Apoio/acoes.csv')

# ===================================================================================================
# Há duas possibilidades de implementação. Buscando em um arquivo csv ou incorporando essas informações
# no próprio script python, conforme implementado a seguir. 
# Analisar qual a melhor abordagem.
# ===================================================================================================
''' Atualizado em 31/01/2022'''
acoes = pd.DataFrame(data=(
    ['3M','MMMC34'],
    ['3MDRN','MMMC34'],
    ['3R PETROLEUMON','RRRP3'],
    ['3TENTOSON','TTEN3'],
    ['524 PARTICIPON','QVQP3B'],
    ['AB INBEV','ABUD34'],
    ['AB INBEVDRN','ABUD34'],
    ['ABB LTD','A1BB34'],
    ['ABB LTDDRN','A1BB34'],
    ['ABBOTT','ABTT34'],
    ['ABBOTTDRN','ABTT34'],
    ['ABBVIE','ABBV34'],
    ['ABBVIEDRN','ABBV34'],
    ['ABC BRASILPN','ABCB4'],
    ['ABIOMED INC','A1BM34'],
    ['ABIOMED INCDRN','A1BM34'],
    ['ACCENTURE','ACNB34'],
    ['ACCENTUREDRN','ACNB34'],
    ['ACO ALTONAON','EALT3'],
    ['ACO ALTONAPN','EALT4'],
    ['ACTIVISION','ATVI34'],
    ['ACTIVISIONDRN','ATVI34'],
    ['ADOBE INC','ADBE34'],
    ['ADOBE INCDRN','ADBE34'],
    ['ADVANCE AUTO','A1AP34'],
    ['ADVANCE AUTODRN','A1AP34'],
    ['ADVANCED MIC','A1MD34'],
    ['ADVANCED MICDRN','A1MD34'],
    ['ADVANCED-DHON','ADHM3'],
    ['AEGON NV','A1EG34'],
    ['AEGON NVDRN','A1EG34'],
    ['AERISON','AERI3'],
    ['AES BRASILON','AESB3'],
    ['AES CORP','A1ES34'],
    ['AES CORPDRN','A1ES34'],
    ['AES TIETE EON','TIET3'],
    ['AES TIETE EPN','TIET4'],
    ['AES TIETE EUNT','TIET11'],
    ['Aesoperacoes ON','AESO3'],
    ['AFLUENTE TON','AFLT3'],
    ['AGRIBRASILON','GRAO3'],
    ['AGROGALAXYON','AGXY3'],
    ['AIRBNB','AIRB34'],
    ['AKAMAI TECHN','A1KA34'],
    ['AKAMAI TECHNDRN','A1KA34'],
    ['ALBEMARLE CO','A1LB34'],
    ['ALBEMARLE CODRN','A1LB34'],
    ['ALEF S/AON','ALEF3B'],
    ['ALEXANDRIA R','A1RE34'],
    ['ALEXANDRIA RDRN','A1RE34'],
    ['ALFA CONSORCON','BRGE3'],
    ['ALFA CONSORCPNA','BRGE5'],
    ['ALFA CONSORCPNB','BRGE6'],
    ['ALFA CONSORCPNC','BRGE7'],
    ['ALFA CONSORCPND','BRGE8'],
    ['ALFA CONSORCPNE','BRGE11'],
    ['ALFA CONSORCPNF','BRGE12'],
    ['ALFA FINANCON','CRIV3'],
    ['ALFA FINANCPN','CRIV4'],
    ['ALFA HOLDINGON','RPAD3'],
    ['ALFA HOLDINGPNA','RPAD5'],
    ['ALFA HOLDINGPNB','RPAD6'],
    ['ALFA INVESTON','BRIV3'],
    ['ALFA INVESTPN','BRIV4'],
    ['ALIANSCSONAEON','ALSO3'],
    ['ALIBABAGR','BABA34'],
    ['ALIBABAGRDRN','BABA34'],
    ['Align Techno DRN','A1LG34'],
    ['Align Techno','A1LG34'],
    ['ALIPERTION','APTI3'],
    ['ALIPERTIPN','APTI4'],
    ['ALL NORTEON','FRRN3B'],
    ['ALL NORTEPNA','FRRN5B'],
    ['ALL NORTEPNB','FRRN6B'],
    ['ALLIANCE DAT','A1LL34'],
    ['ALLIANCE DATDRN','A1LL34'],
    ['ALLIARON','AALR3'],
    ['ALLIEDON','ALLD3'],
    ['ALNYLAM PHAR','A1LN34'],
    ['ALNYLAM PHARDRN','A1LN34'],
    ['ALPARGATASON','ALPA3'],
    ['ALPARGATASPN','ALPA4'],
    ['ALPER S.A.ON','APER3'],
    ['ALPHABET A','GOGL34'],
    ['ALPHABET C','GOGL35'],
    ['ALPHABET','GOGL34'],
    ['ALPHABETDRN A','GOGL34'],
    ['ALPHABETDRN C','GOGL35'],
    ['ALPHAVILLEON','AVLL3'],
    ['ALTERYX INC','A1YX34'],
    ['ALTERYX INCDRN','A1YX34'],
    ['ALTRIA GROUP','MOOO34'],
    ['ALTRIA GROUPDRN','MOOO34'],
    ['ALUPARON','ALUP3'],
    ['ALUPARPN','ALUP4'],
    ['ALUPARUNT','ALUP11'],
    ['AMAZON','AMZO34'],
    ['AMAZONDRN','AMZO34'],
    ['AMAZONIAON','BAZA3'],
    ['AMBEV S/AON','ABEV3'],
    ['AMBIPARON','AMBP3'],
    ['Amcor Plc DRN','A1CR34'],
    ['Amcor Plc','A1CR34'],
    ['AMERICAMOVIL','A1MX34'],
    ['AMERICAMOVILDRN','A1MX34'],
    ['AMERICAN AIR','AALL34'],
    ['AMERICAN AIRDRN','AALL34'],
    ['AMERICAN EXP','AXPB34'],
    ['AMERICAN EXPDRN','AXPB34'],
    ['AMERICAN TOW','T1OW34'],
    ['AMERICAN TOWDRN','T1OW34'],
    ['AMERICANASON','AMER3'],
    ['AMGEN','AMGN34'],
    ['AMGENDRN','AMGN34'],
    ['AMPLA ENERGON','CBEE3'],
    ['ANGLOGOLD AS','A1UA34'],
    ['ANGLOGOLD ASDRN','A1UA34'],
    ['ANIMAON','ANIM3'],
    ['APA CORP','A1PA34'],
    ['APA CORPDRN','A1PA34'],
    ['APARTMENT IN','A1IV34'],
    ['APARTMENT INDRN','A1IV34'],
    ['APPLE','AAPL34'],
    ['APPLEDRN','AAPL34'],
    ['APPLIEDTE','A1MT34'],
    ['APPLIEDTEDRN','A1MT34'],
    ['ARCELOR','ARMT34'],
    ['ARCELORDRN','ARMT34'],
    ['AREZZO COON','ARZZ3'],
    ['ARMACON','ARML3'],
    ['ASML HOLD','ASML34'],
    ['ASML HOLDDRN','ASML34'],
    ['ASSAION','ASAI3'],
    ['ASTRAZENECA','A1ZN34'],
    ['ASTRAZENECADRN','A1ZN34'],
    ['Atlassian Co DRN','T1AM34'],
    ['Atlassian Co','T1AM34'],
    ['ATMASAON','ATMP3'],
    ['ATOMPARON','ATOM3'],
    ['ATT INC','ATTB34'],
    ['ATT INCDRN','ATTB34'],
    ['AURA 360ON','AURA33'],
    ['Auren ON NM','AURE3'],
    ['AUTODESK INC','A1UT34'],
    ['AUTODESK INCDRN','A1UT34'],
    ['Autohome Inc DRN','A1TH34'],
    ['Autohome Inc','A1TH34'],
    ['AVALONBAY CO','A1VB34'],
    ['AVALONBAY CODRN','A1VB34'],
    ['AZEVEDOON','AZEV3'],
    ['AZEVEDOPN','AZEV4'],
    ['AZUL S.A.PN','AZUL4'],
    ['AZULPN','AZUL4'],
    ['B TECH EQION','BLUT3'],
    ['B TECH EQION','JBDU3'],
    ['B TECH EQIPN','BLUT4'],
    ['B TECH EQIPN','JBDU4'],
    ['B2W - COMPANON','AMER3'],
    ['B2W DIGITALON','AMER3'],
    ['B3ON','B3SA3'],
    ['BAHEMAON','BAHI3'],
    ['BAIDU INC','BIDU34'],
    ['BAIDU INCDRN','BIDU34'],
    ['BANCO BMGPN','BMGB4'],
    ['BANCO INTERON','BIDI3'],
    ['BANCO INTERPN','BIDI4'],
    ['BANCO INTERUNT','BIDI11'],
    ['BANCO PANPN','BPAN4'],
    ['BANCO SANTAN','B1SA34'],
    ['BANCO SANTANDRN','B1SA34'],
    ['BANESEON','BGIP3'],
    ['BANESEPN','BGIP4'],
    ['BANESTESON','BEES3'],
    ['BANESTESPN','BEES4'],
    ['BANK AMERICA','BOAC34'],
    ['BANK AMERICADRN','BOAC34'],
    ['BANPARAON','BPAR3'],
    ['BANRISULON','BRSR3'],
    ['BANRISULPNA','BRSR5'],
    ['BANRISULPNB','BRSR6'],
    ['BARCLAYS PLC','B1CS34'],
    ['BARCLAYS PLCDRN','B1CS34'],
    ['BARDELLAON','BDLL3'],
    ['BARDELLAPN','BDLL4'],
    ['BATTISTELLAON','EPAR3'],
    ['BAUMERON','BALM3'],
    ['BAUMERPN','BALM4'],
    ['BB ETF IAGRO','AGRI11'],
    ['BB ETF IAGROCI','AGRI11'],
    ['BB ETF IBOV','BBOV11'],
    ['BB ETF IBOVCI','BBOV11'],
    ['BB ETF MILHO','CORN11'],
    ['BB ETF MILHOCI','CORN11'],
    ['BB ETF SP DV','BBSD11'],
    ['BB ETF SP DVCI','BBSD11'],
    ['BBMLOGISTICAON','BBML3'],
    ['BBSEGURIDADEON','BBSE3'],
    ['BCO BRADESCOPN','BBDC4'],
    ['BCO BRASIL SON','BBAS3'],
    ['BCO ESTADO DPNB','BRSR6'],
    ['BEMOBI TECHON','BMOB3'],
    ['BERKSHIRE','BERK34'],
    ['BERKSHIREDRN','BERK34'],
    ['BEST BUY','BBYY34'],
    ['BEST BUYDRN','BBYY34'],
    ['BETAPARTON','BETP3B'],
    ['Beyond Meat DRN','B2YN34'],
    ['Beyond Meat','B2YN34'],
    ['Bhp Group DRN','BHPG34'],
    ['Bhp Group Pl DRN','B1BL34'],
    ['Bhp Group Pl','B1BL34'],
    ['Bhp Group','BHPG34'],
    ['BIC MONARKON','BMKS3'],
    ['BILIBILI INC','B1IL34'],
    ['BILIBILI INCDRN','B1IL34'],
    ['BIOGEN','BIIB34'],
    ['BIOGENDRN','BIIB34'],
    ['BIOMMON','BIOM3'],
    ['BIONTECH SE','B1NT34'],
    ['BIONTECH SEDRN','B1NT34'],
    ['BIOSEVON','BSEV3'],
    ['BITCOIN HASH','BITH11'],
    ['BITCOIN HASHCI','BITH11'],
    ['BK BRASILON','BKBR3'],
    ['BLACKROCK','BLAK34'],
    ['BLACKROCKDRN','BLAK34'],
    ['BLAUON','BLAU3'],
    ['Block Inc. DRN','S2QU34'],
    ['Block Inc.','S2QU34'],
    ['BMFBOVESPAON','B3SA3'],
    ['BNY MELLON','BONY34'],
    ['BNY MELLONDRN','BONY34'],
    ['BOA SAFRAON','SOJA3'],
    ['BOA VISTAON','BOAS3'],
    ['BOEING','BOEI34'],
    ['BOEINGDRN','BOEI34'],
    ['BOMBRILON','BOBR3'],
    ['BOMBRILPN','BOBR4'],
    ['BOOKING','BKNG34'],
    ['BOOKINGDRN','BKNG34'],
    ['BOSTON PROP','BOXP34'],
    ['BOSTON PROPDRN','BOXP34'],
    ['BOSTON SCIEN','B1SX34'],
    ['BOSTON SCIENDRN','B1SX34'],
    ['BP PLC','B1PP34'],
    ['BP PLCDRN','B1PP34'],
    ['BR BROKERSON','BBRK3'],
    ['BR MALLS PARON','BRML3'],
    ['Br Partners UNT N2','BRBI11'],
    ['BR PROPERTON','BRPR3'],
    ['BRADESCOON','BBDC3'],
    ['BRADESCOPN','BBDC4'],
    ['BRADESPARON','BRAP3'],
    ['BRADESPARPN','BRAP4'],
    ['BRASILAGROON','AGRO3'],
    ['BRASILON','BBAS3'],
    ['BRASKEMON','BRKM3'],
    ['BRASKEMPNA','BRKM5'],
    ['BRASKEMPNB','BRKM6'],
    ['BRB BANCOON','BSLI3'],
    ['BRB BANCOPN','BSLI4'],
    ['BRF S.A.ON','BRFS3'],
    ['BRF SAON','BRFS3'],
    ['BRISANETON','BRIT3'],
    ['BRISTOLMYERS','BMYB34'],
    ['BRISTOLMYERSDRN','BMYB34'],
    ['BRITISH AMER','B1TI34'],
    ['BRITISH AMERDRN','B1TI34'],
    ['BRLLS PARON','BRML3'],
    ['BROADCOM INC','AVGO34'],
    ['BROADCOM INCDRN','AVGO34'],
    ['BRQON','BRQB3'],
    ['BTG COMMODIT','CMDB11'],
    ['BTG COMMODITCI','CMDB11'],
    ['BTG S&P 500','SPXB11'],
    ['BTG S&P 500CI','SPXB11'],
    ['BTG SMLL CAP','SMAB11'],
    ['BTG SMLL CAPCI','SMAB11'],
    ['BTGP BANCOON','BPAC3'],
    ['BTGP BANCOPNA','BPAC5'],
    ['BTGP BANCOUNT','BPAC11'],
    ['CABINDA PARTON','CABI3B'],
    ['CACONDE PARTON','CACO3B'],
    ['CAIXA SEGURION','CXSE3'],
    ['CAIXAETFXBOV','XBOV11'],
    ['CAIXAETFXBOVCI','XBOV11'],
    ['CAMBUCION','CAMB3'],
    ['CAMILON','CAML3'],
    ['CAPRI HOLDI','CAPH34'],
    ['CAPRI HOLDIDRN','CAPH34'],
    ['CARNIVAL COR','C1CL34'],
    ['CARNIVAL CORDRN','C1CL34'],
    ['CARREFOUR BRON','CRFB3'],
    ['CASANON','CASN3'],
    ['CASANPN','CASN4'],
    ['CATERPILLAR','CATP34'],
    ['CATERPILLARDRN','CATP34'],
    ['CBAON','CBAV3'],
    ['CCR S.A.ON ED','CCRO3'],
    ['CCR S.A.ON','CCRO3'],
    ['CCR SAON','CCRO3'],
    ['CEA MODASON','CEAB3'],
    ['CEBON','CEBR3'],
    ['CEBPNA','CEBR5'],
    ['CEBPNB','CEBR6'],
    ['CEDROON','CEDO3'],
    ['CEDROPN','CEDO4'],
    ['CEEE-DON','CEED3'],
    ['CEEE-DPN','CEED4'],
    ['CEEE-GTON','EEEL3'],
    ['CEEE-GTPN','EEEL4'],
    ['CEEE-TON','EEEL3'],
    ['CEEE-TPN','EEEL4'],
    ['CEGON','CEGR3'],
    ['CELESCON','CLSC3'],
    ['CELESCPN','CLSC4'],
    ['CELGPARON','GPAR3'],
    ['CELPEON','CEPE3'],
    ['CELPEPNA','CEPE5'],
    ['CELPEPNB','CEPE6'],
    ['CELUL IRANIPN','RANI4'],
    ['CEMEPEON','MAPT3'],
    ['CEMEPEPN','MAPT4'],
    ['CEMIGON','CMIG3'],
    ['CEMIGPN','CMIG4'],
    ['CENTAUROON','SBFG3'],
    ['CENTRAIS ELEON','ELET3'],
    ['CESPON','CESP3'],
    ['CESPPNA','CESP5'],
    ['CESPPNB','CESP6'],
    ['CHARTER COMM','CHCM34'],
    ['CHARTER COMMDRN','CHCM34'],
    ['CHEVRON','CHVX34'],
    ['CHEVRONDRN','CHVX34'],
    ['China Life I DRN','L1FC34'],
    ['China Life I','L1FC34'],
    ['CHINA PETROL','C1HI34'],
    ['CHINA PETROLDRN','C1HI34'],
    ['CIA FERRO LIPN','FESA4'],
    ['CIA HERINGON','HGTX3'],
    ['CIA LOCAÇÃOON','LCAM3'],
    ['CIA SANEAMENON','SAPR3'],
    ['CIA SIDERURGON','CSNA3'],
    ['CIELOON','CIEL3'],
    ['CIGNA CORP','C1IC34'],
    ['CIGNA CORPDRN','C1IC34'],
    ['CIMSON','CMSA3'],
    ['CIMSPN','CMSA4'],
    ['CINESYSTEMON','CNSY3'],
    ['CISCO','CSCO34'],
    ['CISCODRN','CSCO34'],
    ['CITIGROUP','CTGP34'],
    ['CITIGROUPDRN','CTGP34'],
    ['CLEARSALEON','CLSA3'],
    ['COCA COLA','COCA34'],
    ['COCA COLADRN','COCA34'],
    ['COELBAON','CEEB3'],
    ['COELBAPNA','CEEB5'],
    ['COELBAPNB','CEEB6'],
    ['COELCEON','COCE3'],
    ['COELCEPNA','COCE5'],
    ['COELCEPNB','COCE6'],
    ['COGNA ONON','COGN3'],
    ['COGNAONON','COGN3'],
    ['COLGATE','COLG34'],
    ['COLGATEDRN','COLG34'],
    ['COMCAST','CMCS34'],
    ['COMCASTDRN','CMCS34'],
    ['Comerc Par ON','COMR3'],
    ['COMGASON','CGAS3'],
    ['COMGASPNA','CGAS5'],
    ['COMPASSON','PASS3'],
    ['COMPASSPNA','PASS5'],
    ['COMPASSPNB','PASS6'],
    ['CONC RIO TERON','CRTE3B'],
    ['CONC RIO TERPNA','CRTE5B'],
    ['CONST A LINDON','CALI3'],
    ['CONST A LINDPN','CALI4'],
    ['COPASAON','CSMG3'],
    ['Copel UNT N2','CPLE11'],
    ['COPELON','CPLE3'],
    ['COPELPNA','CPLE5'],
    ['COPELPNB','CPLE6'],
    ['COPHILLIPS','COPH34'],
    ['COPHILLIPSDRN','COPH34'],
    ['COR RIBEIROON','CORR3'],
    ['COR RIBEIROPN','CORR4'],
    ['CORTEVA INC','C1TV34'],
    ['CORTEVA INCDRN','C1TV34'],
    ['COSAN LOGON','RLOG3'],
    ['COSANON','CSAN3'],
    ['COSERNON','CSRN3'],
    ['COSERNPNA','CSRN5'],
    ['COSERNPNB','CSRN6'],
    ['Costar Group DRN','C1GP34'],
    ['Costar Group','C1GP34'],
    ['COSTCO','COWC34'],
    ['COSTCODRN','COWC34'],
    ['COTEMINASON','CTNM3'],
    ['COTEMINASPN','CTNM4'],
    ['COTY INC','COTY34'],
    ['COTY INCDRN','COTY34'],
    ['Coupa Softwa DRN','C1OU34'],
    ['Coupa Softwa','C1OU34'],
    ['CPFL ENERGIAON ED','CPFE3'],
    ['CPFL ENERGIAON','CPFE3'],
    ['CR2ON','CRDE3'],
    ['CREDIT SUISS','C1SU34'],
    ['CREDIT SUISSDRN','C1SU34'],
    ['CRIPTO20 EMP','CRPT11'],
    ['CRIPTO20 EMPCI','CRPT11'],
    ['CRISTALON','CRPG3'],
    ['CRISTALPNA','CRPG5'],
    ['CRISTALPNB','CRPG6'],
    ['CRUZEIRO EDUON','CSED3'],
    ['CRUZEIROUON','CSED3'],
    ['CSINERACAOON','CMIN3'],
    ['CSN MINERACAON','CMIN3'],
    ['CSU CARDSYSTON','CARD3'],
    ['CTC S.A.ON','CTCA3'],
    ['CTRIPCOM','CRIP34'],
    ['CTRIPCOMDRN','CRIP34'],
    ['CURY S/AON','CURY3'],
    ['CVC BRASILON','CVCB3'],
    ['CVS HEALTH','CVSH34'],
    ['CVS HEALTHDRN','CVSH34'],
    ['CYRE COM-CCPON','CCPR3'],
    ['CYRELA REALTON','CYRE3'],
    ['D1000VFARMA','DMVF3'],
    ['D1000VFARMAON','DMVF3'],
    ['DANAHER CORP','DHER34'],
    ['DANAHER CORPDRN','DHER34'],
    ['DASAON','DASA3'],
    ['Datadog Inc DRN','D1DG34'],
    ['Datadog Inc','D1DG34'],
    ['DEERE CO','DEEC34'],
    ['DEERE CODRN','DEEC34'],
    ['DEFI HASH','DEFI11'],
    ['DEFI HASHCI','DEFI11'],
    ['DELL TECHNOL','D1EL34'],
    ['DELL TECHNOLDRN','D1EL34'],
    ['DELTA','DEAI34'],
    ['DELTADRN','DEAI34'],
    ['DESKTOPON','DESK3'],
    ['DEUTSCHE AK','DBAG34'],
    ['DEUTSCHE AKDRN','DBAG34'],
    ['Devon Energy DRN','D1VN34'],
    ['Devon Energy','D1VN34'],
    ['Dexcom Inc DRN','D1EX34'],
    ['Dexcom Inc','D1EX34'],
    ['DEXCOON','DXCO3'],
    ['DEXXOS PARON','DEXP3'],
    ['DEXXOS PARPN','DEXP4'],
    ['Diageo Pl DRN','DEOP34'],
    ['Diageo Pl','DEOP34'],
    ['DIAMONDBACK','F1AN34'],
    ['DIAMONDBACKDRN','F1AN34'],
    ['DIGITAL REAL','D1LR34'],
    ['DIGITAL REALDRN','D1LR34'],
    ['DIMEDON','PNVL3'],
    ['DIMEDPN','PNVL4'],
    ['DIRECIONALON','DIRR3'],
    ['Discovery In ','DCVY34'],
    ['Discovery In DRN A','DCVY34'],
    ['DOCUSIGN INC','D1OC34'],
    ['DOCUSIGN INCDRN','D1OC34'],
    ['DOHLERON','DOHL3'],
    ['DOHLERPN','DOHL4'],
    ['DOLLAR GENER','DGCO34'],
    ['DOLLAR GENERDRN','DGCO34'],
    ['DOMMOON','DMMO3'],
    ['DOTZ SAON','DOTZ3'],
    ['Dow Inc DRN','D1OW34'],
    ['Dow Inc','D1OW34'],
    ['DTCOM DIRECTON','DTCY3'],
    ['DTCOM DIRECTPN','DTCY4'],
    ['DTCOM-DIRECTON','DTCY3'],
    ['DTCOM-DIRECTPN','DTCY4'],
    ['DURATEXON','DTEX3'],
    ['Eaton Corp P DRN','E1TN34'],
    ['Eaton Corp P','E1TN34'],
    ['EBAY','EBAY34'],
    ['EBAYDRN','EBAY34'],
    ['ECORODOVIASON','ECOR3'],
    ['ELECTR ARTS','EAIN34'],
    ['ELECTR ARTSDRN','EAIN34'],
    ['ELEKTROON','EKTR3'],
    ['ELEKTROPN','EKTR4'],
    ['ELETROBRASON','ELET3'],
    ['ELETROBRASPNA','ELET5'],
    ['ELETROBRASPNB','ELET6'],
    ['ELETROMIDIAON','ELMD3'],
    ['ELETROPARON','LIPR3'],
    ['ELETROPAULOPN','ELPL4'],
    ['EMAEON','EMAE3'],
    ['EMAEPN','EMAE4'],
    ['EMBPAR S/AON','EPAR3'],
    ['EMBRAERON','EMBR3'],
    ['EMBRATEL PARON','EBTP3'],
    ['EMBRATEL PARPN','EBTP4'],
    ['ENAUTA PARTON','ENAT3'],
    ['ENCORPARON','ECPR3'],
    ['ENCORPARPN','ECPR4'],
    ['ENEL AMERICA','E1NI34'],
    ['ENEL AMERICADRN','E1NI34'],
    ['ENERGIAS BRON','ENBR3'],
    ['ENERGISA MTON','ENMT3'],
    ['ENERGISA MTPN','ENMT4'],
    ['ENERGISAON','ENGI3'],
    ['ENERGISAPN','ENGI4'],
    ['ENERGISAUNT','ENGI11'],
    ['ENEVAON','ENEV3'],
    ['ENGIE BRASILON','EGIE3'],
    ['ENJOEION','ENJU3'],
    ['EQTL MARANHAOON','EQMA3B'],
    ['EQTL MARANHAOPNA','EQMA5B'],
    ['EQTL MARANHAOPNB','EQMA6B'],
    ['EQTL PARAON','EQPA3'],
    ['EQTL PARAPNA','EQPA5'],
    ['EQTL PARAPNB','EQPA6'],
    ['EQTL PARAPNC','EQPA7'],
    ['EQUATORIALON','EQTL3'],
    ['EQUINIX INC','EQIX34'],
    ['EQUINIX INCDRN','EQIX34'],
    ['Equinor Asa DRN','E1QN34'],
    ['Equinor Asa','E1QN34'],
    ['EQUITY RESID','E1QR34'],
    ['EQUITY RESIDDRN','E1QR34'],
    ['ERICSSON LM','E1RI34'],
    ['ERICSSON LMDRN','E1RI34'],
    ['ESPACOLASERON','ESPA3'],
    ['ESSEX PROPER','E1SS34'],
    ['ESSEX PROPERDRN','E1SS34'],
    ['ESTACIO PARTON','YDUQ3'],
    ['ESTAPARON','ALPK3'],
    ['Estee Lauder DRN','ELCI34'],
    ['Estee Lauder','ELCI34'],
    ['ESTRELAON','ESTR3'],
    ['ESTRELAPN','ESTR4'],
    ['ETERNITON','ETER3'],
    ['ETF BRA IBOV','BOVB11'],
    ['ETF BRA IBOVCI','BOVB11'],
    ['ETF BTG GENB','GENB11'],
    ['ETF BTG GENBCI','GENB11'],
    ['ETF ESG BTG','ESGB11'],
    ['ETF ESG BTGCI','ESGB11'],
    ['ETF GURU','GURU11'],
    ['ETF GURUCI','GURU11'],
    ['ETHER HASH','ETHE11'],
    ['ETHER HASHCI','ETHE11'],
    ['EUCATEXON','EUCA3'],
    ['EUCATEXPN','EUCA4'],
    ['EVENON','EVEN3'],
    ['EXCELSIORON','BAUH3'],
    ['EXCELSIORPN','BAUH4'],
    ['EXXON MOBIL','EXXO34'],
    ['EXXON MOBILDRN','EXXO34'],
    ['EZ TEC EMPREON','EZTC3'],
    ['EZTECON','EZTC3'],
    ['FACEBOOK','FBOK34'],
    ['FACEBOOKDRN','FBOK34'],
    ['FEDERAL REAL','F1RI34'],
    ['FEDERAL REALDRN','F1RI34'],
    ['FEDEX CORP','FDXB34'],
    ['FEDEX CORPDRN','FDXB34'],
    ['FER C ATLANTON','VSPT3'],
    ['FER C ATLANTPN','VSPT4'],
    ['FER HERINGERON','FHER3'],
    ['FERBASAON','FESA3'],
    ['FERBASAPN','FESA4'],
    ['FIBRIA CELULON','FIBR3'],
    ['FIDELITY NAT','F1NI34'],
    ['FIDELITY NATDRN','F1NI34'],
    ['FINANSINOSON','FNCN3'],
    ['FIRST SOLAR','FSLR34'],
    ['FIRST SOLARDRN','FSLR34'],
    ['FLEURYON','FLRY3'],
    ['FLEX S/AON','FLEX3'],
    ['FOCUS ONON','POWE3'],
    ['FOCUSONON','POWE3'],
    ['FORD MOTORS','FDMO34'],
    ['FORD MOTORSDRN','FDMO34'],
    ['FRAS-LEON','FRAS3'],
    ['FREEPORT','FCXO34'],
    ['FREEPORTDRN','FCXO34'],
    ['GAFISAON','GFSA3'],
    ['Galapagos Nv DRN','G1LP34'],
    ['Galapagos Nv','G1LP34'],
    ['GAMA PARTON','OPGM3B'],
    ['GE','GEOO34'],
    ['GEDRN','GEOO34'],
    ['GENERAL MOT','GMCO34'],
    ['GENERAL MOTDRN','GMCO34'],
    ['GENERALSHOPPON','GSHP3'],
    ['GER PARANAPON','GEPA3'],
    ['GER PARANAPPN','GEPA4'],
    ['GERDAU METON','GOAU3'],
    ['GERDAU METPN','GOAU4'],
    ['GERDAUON','GGBR3'],
    ['GERDAUPN','GGBR4'],
    ['Getnet Br UNT','GETT11'],
    ['GETNET BRON','GETT3'],
    ['GETNET BRPN','GETT4'],
    ['GETNINJASON','NINJ3'],
    ['GILEAD','GILD34'],
    ['GILEADDRN','GILD34'],
    ['GLAXOSMITHKL','G1SK34'],
    ['GLAXOSMITHKLDRN','G1SK34'],
    ['Gol ON','GOLL3'],
    ['GOLD FIELDS','G1FI34'],
    ['GOLD FIELDSDRN','G1FI34'],
    ['GOLDMANSACHS','GSGI34'],
    ['GOLDMANSACHSDRN','GSGI34'],
    ['GOLPN','GOLL4'],
    ['GOPRO','GPRO34'],
    ['GOPRODRN','GPRO34'],
    ['GP INVESTA','GPIV33'],
    ['GPC PARTON','DEXP3'],
    ['GPC PARTPN','DEXP4'],
    ['GPSON','GGPS3'],
    ['GRAZZIOTINON','CGRA3'],
    ['GRAZZIOTINPN','CGRA4'],
    ['GRENDENEON','GRND3'],
    ['GRUPO MATEUSON','GMAT3'],
    ['GRUPO NATURA','NTCO3'],
    ['GRUPO NATURAON','NTCO3'],
    ['GRUPO SBFON','SBFG3'],
    ['GRUPO SOMAON','SOMA3'],
    ['GRUPOTEUSON','GMAT3'],
    ['GUARARAPESON','GUAR3'],
    ['HABITASULON','HBTS3'],
    ['HABITASULPNA','HBTS5'],
    ['HABITASULPNB','HBTS6'],
    ['HAGA S/AON','HAGA3'],
    ['HAGA S/APN','HAGA4'],
    ['HAPVIDAON','HAPV3'],
    ['Harley-David DRN','H1OG34'],
    ['Harley-David','H1OG34'],
    ['HASHDEX NCI','HASH11'],
    ['HASHDEX NCICI','HASH11'],
    ['HBR REALTYON','HBRE3'],
    ['Hca Healthca DRN','H1CA34'],
    ['Hca Healthca','H1CA34'],
    ['HELBORON','HBOR3'],
    ['HERCULESON','HETA3'],
    ['HERCULESPN','HETA4'],
    ['HIDROVIASON','HBSA3'],
    ['Hmobi S.A ON','HMOB3'],
    ['HOME DEPOT','HOME34'],
    ['HOME DEPOTDRN','HOME34'],
    ['HONDA MO','HOND34'],
    ['HONDA MODRN','HOND34'],
    ['HORIZON THER','H1ZN34'],
    ['HORIZON THERDRN','H1ZN34'],
    ['HOTEIS OTHONON','HOOT3'],
    ['HOTEIS OTHONPN','HOOT4'],
    ['HP COMPANY','HPQB34'],
    ['HP COMPANYDRN','HPQB34'],
    ['HSBC HOLDING','H1SB34'],
    ['HSBC HOLDINGDRN','H1SB34'],
    ['HYPERAON','HYPE3'],
    ['IBM EC','IBMB34'],
    ['IBMDRN EC','IBMB34'],
    ['IGB S/AON','IGBR3'],
    ['IGUA SAON','IGSN3'],
    ['Iguatemi S.A ON N1','IGTI3'],
    ['Iguatemi S.A UNT N1','IGTI11'],
    ['IGUATEMION','IGTA3'],
    ['IHPARDINION','PARD3'],
    ['ILLUMINA INC','I1LM34'],
    ['ILLUMINA INCDRN','I1LM34'],
    ['IMC S/AON','MEAL3'],
    ['IND CATAGUASON','CATA3'],
    ['IND CATAGUASPN','CATA4'],
    ['INDS ROMION','ROMI3'],
    ['INDUSTRIAS RON','ROMI3'],
    ['INDUSVALON','IDVL3'],
    ['INDUSVALPN','IDVL4'],
    ['INEPARON','INEP3'],
    ['INEPARPN','INEP4'],
    ['INFRACOMMON','IFCM3'],
    ['INTEL','ITLC34'],
    ['INTELBRASON','INTB3'],
    ['INTELDRN','ITLC34'],
    ['INTER SAON','INNT3'],
    ['INTERMEDICAON','GNDI3'],
    ['Intuit Inc DRN','INTU34'],
    ['Intuit Inc','INTU34'],
    ['Intuitive Su DRN','I1SR34'],
    ['Intuitive Su','I1SR34'],
    ['INVEPARON','IVPR3B'],
    ['INVEPARPN','IVPR4B'],
    ['INVEST BEMGEON','FIGE3'],
    ['INVEST BEMGEPN','FIGE4'],
    ['INVESTO 5GTK','5GTK11'],
    ['INVESTO 5GTKCI','5GTK11'],
    ['INVESTO ALUG','ALUG11'],
    ['INVESTO ALUGCI','ALUG11'],
    ['INVESTO BDOM','BDOM11'],
    ['INVESTO BDOMCI','BDOM11'],
    ['INVESTO BLOK','BLOK11'],
    ['INVESTO BLOKCI','BLOK11'],
    ['INVESTO BTEK','BTEK11'],
    ['INVESTO BTEKCI','BTEK11'],
    ['INVESTO BXPO','BXPO11'],
    ['INVESTO BXPOCI','BXPO11'],
    ['INVESTO FOOD','FOOD11'],
    ['INVESTO FOODCI','FOOD11'],
    ['INVESTO JOGO','JOGO11'],
    ['INVESTO JOGOCI','JOGO11'],
    ['INVESTO NFTS','NFTS11'],
    ['INVESTO NFTSCI','NFTS11'],
    ['INVESTO PEVC','PEVC11'],
    ['INVESTO PEVCCI','PEVC11'],
    ['INVESTO SCVB','SCVB11'],
    ['INVESTO SCVBCI','SCVB11'],
    ['INVESTO SVAL','SVAL11'],
    ['INVESTO SVALCI','SVAL11'],
    ['INVESTO USTK','USTK11'],
    ['INVESTO USTKCI','USTK11'],
    ['INVESTO WRLD','WRLD11'],
    ['INVESTO WRLDCI','WRLD11'],
    ['IOCHP-MAXIONON','MYPK3'],
    ['IRANION','RANI3'],
    ['IRBBRASIL REON','IRBR3'],
    ['IRON MOUNTAI','I1RM34'],
    ['IRON MOUNTAIDRN','I1RM34'],
    ['ISHARE SP500','IVVB11'],
    ['ISHARE SP500CI','IVVB11'],
    ['ISHARES BOVA','BOVA11'],
    ['ISHARES BOVACI','BOVA11'],
    ['ISHARES BOVACI','BOVA11'],
    ['ISHARES BRAX','BRAX11'],
    ['ISHARES BRAXCI','BRAX11'],
    ['ISHARES ECOO','ECOO11'],
    ['ISHARES ECOOCI','ECOO11'],
    ['ISHARES SMAL','SMAL11'],
    ['ISHARES SMALCI','SMAL11'],
    ['IT NOW DNA','DNAI11'],
    ['IT NOW DNACI','DNAI11'],
    ['IT NOW GREEN','REVE11'],
    ['IT NOW GREENCI','REVE11'],
    ['IT NOW HCARE','HTEK11'],
    ['IT NOW HCARECI','HTEK11'],
    ['IT NOW HYDRO','YDRO11'],
    ['IT NOW HYDROCI','YDRO11'],
    ['IT NOW IBOV','BOVV11'],
    ['IT NOW IBOVCI','BOVV11'],
    ['IT NOW IDIV','DIVO11'],
    ['IT NOW IDIVCI','DIVO11'],
    ['IT NOW IFNC','FIND11'],
    ['IT NOW IFNCCI','FIND11'],
    ['IT NOW IGCT','GOVE11'],
    ['IT NOW IGCTCI','GOVE11'],
    ['IT NOW IMAT','MATB11'],
    ['IT NOW IMATCI','MATB11'],
    ['IT NOW ISE','ISUS11'],
    ['IT NOW ISECI','ISUS11'],
    ['IT NOW MILL','MILL11'],
    ['IT NOW MILLCI','MILL11'],
    ['IT NOW PIBB','PIBB11'],
    ['IT NOW PIBBCI','PIBB11'],
    ['IT NOW SHOT','SHOT11'],
    ['IT NOW SHOTCI','SHOT11'],
    ['IT NOW SMALL','SMAC11'],
    ['IT NOW SMALLCI','SMAC11'],
    ['IT NOW SPXI','SPXI11'],
    ['IT NOW SPXICI','SPXI11'],
    ['IT NOW TECK','TECK11'],
    ['IT NOW TECKCI','TECK11'],
    ['ITAUBANCOPN','ITUB4'],
    ['ITAUSAON','ITSA3'],
    ['ITAUSAPN','ITSA4'],
    ['ITAUUNIBANCOON','ITUB3'],
    ['ITAUUNIBANCOPN','ITUB4'],
    ['J B DUARTEON','JBDU3'],
    ['J B DUARTEPN','JBDU4'],
    ['JALLESMACHADON','JALL3'],
    ['JBSON','JBSS3'],
    ['JD COM','JDCO34'],
    ['JD COMDRN','JDCO34'],
    ['JEREISSATI PPN','JPSA3'],
    ['JEREISSATION','JPSA3'],
    ['JHSF PARTON','JHSF3'],
    ['JOAO FORTESON','JFEN3'],
    ['JOHNSON','JNJB34'],
    ['JOHNSONDRN','JNJB34'],
    ['JOSAPARON','JOPA3'],
    ['JOSAPARPN','JOPA4'],
    ['JPMORGAN','JPMC34'],
    ['JPMORGANDRN','JPMC34'],
    ['JSLON','JSLG3'],
    ['KALLASON','KLAS3'],
    ['KARSTENON','CTKA3'],
    ['KARSTENPN','CTKA4'],
    ['KEPLER WEBERON','KEPL3'],
    ['KLABIN S.A.UNT','KLBN11'],
    ['KLABIN S/AON','KLBN3'],
    ['KLABIN S/APN','KLBN4'],
    ['KLABIN S/AUNT','KLBN11'],
    ['KORA SAUDEON','KRSA3'],
    ['KRAFT HEINZ','KHCB34'],
    ['KRAFT HEINZDRN','KHCB34'],
    ['KROTON EDUCAON','COGN3'],
    ['KROTONON','COGN3'],
    ['KROTONUCAON','COGN3'],
    ['LAM RESEARCH','L1RC34'],
    ['LAM RESEARCHDRN','L1RC34'],
    ['LAVVION','LAVV3'],
    ['LE BISCUITON','LLBI3'],
    ['LE BISCUITPN','LLBI4'],
    ['LE LIS BLANCON','LLIS3'],
    ['LIFEMEDON','LMED3'],
    ['LIGHT S/AON','LIGT3'],
    ['LILLY','LILY34'],
    ['LILLYDRN','LILY34'],
    ['LINXON','LINX3'],
    ['LITELAON','LTLA3B'],
    ['LITELON','LTEL3B'],
    ['LLOYDS BANKI','L1YG34'],
    ['LLOYDS BANKIDRN','L1YG34'],
    ['LOCALIZAON','RENT3'],
    ['LOCAMERICAON','LCAM3'],
    ['LOCAWEBON','LWSA3'],
    ['Lockheed DRN','LMTB34'],
    ['Lockheed','LMTB34'],
    ['LOG COM PROPON','LOGG3'],
    ['LOG-INON','LOGN3'],
    ['LOJAS AMERICON','LAME3'],
    ['LOJAS AMERICPN','LAME4'],
    ['LOJAS MARISAON','AMAR3'],
    ['LOJAS RENNERON','LREN3'],
    ['LOJASRISAON','AMAR3'],
    ['LOPES BRASILON','LPSB3'],
    ['LOWES COMPA','LOWC34'],
    ['LOWES COMPADRN','LOWC34'],
    ['Lululemon At DRN','L1UL34'],
    ['Lululemon At','L1UL34'],
    ['LUMEN TECH','L1MN34'],
    ['LUMEN TECHDRN','L1MN34'],
    ['LUPATECHON','LUPA3'],
    ['M.DIASBRANCOON','MDIA3'],
    ['MACY S','MACY34'],
    ['MACY SDRN','MACY34'],
    ['MAESTROLOCON','MSRO3'],
    ['MAGAZ LUIZAON','MGLU3'],
    ['MAGAZINE LUION','MGLU3'],
    ['MANGELS INDLON','MGEL3'],
    ['MANGELS INDLPN','MGEL4'],
    ['Marathon Oil DRN','M1RO34'],
    ['Marathon Oil','M1RO34'],
    ['MARCOPOLO S.ON','POMO3'],
    ['MARCOPOLO S.PN','POMO4'],
    ['MARCOPOLOON','POMO3'],
    ['MARCOPOLOPN','POMO4'],
    ['MARFRIGON','MRFG3'],
    ['MASTERCARD','MSCD34'],
    ['MASTERCARDDRN','MSCD34'],
    ['MATER DEION','MATD3'],
    ['MATERDEION','MATD3'],
    ['MCDONALDS','MCDC34'],
    ['MCDONALDSDRN','MCDC34'],
    ['MEDTRONIC','MDTC34'],
    ['MEDTRONICDRN','MDTC34'],
    ['MELHOR SPON','MSPA3'],
    ['MELHOR SPPN','MSPA4'],
    ['MELIUZON','CASH3'],
    ['MELNICKON','MELK3'],
    ['MENEZES CORTON','MNZC3B'],
    ['MERC BRASILON','BMEB3'],
    ['MERC BRASILPN','BMEB4'],
    ['MERC FINANCON','MERC3'],
    ['MERC FINANCPN','MERC4'],
    ['MERC INVESTON','BMIN3'],
    ['MERC INVESTPN','BMIN4'],
    ['MERCADOLIBRE','MELI34'],
    ['MERCADOLIBREDRN','MELI34'],
    ['META HASH','META11'],
    ['META HASHCI','META11'],
    ['METAL IGUACUON','MTIG3'],
    ['METAL IGUACUPN','MTIG4'],
    ['METAL LEVEON','LEVE3'],
    ['METALFRIOON','FRIO3'],
    ['METISAON','MTSA3'],
    ['METISAPN','MTSA4'],
    ['Metlife Inc DRN','METB34'],
    ['Metlife Inc','METB34'],
    ['MGM RESORTS','M1GM34'],
    ['MGM RESORTSDRN','M1GM34'],
    ['MICRON TECHN','MUTC34'],
    ['MICRON TECHNDRN','MUTC34'],
    ['MICROSOFT','MSFT34'],
    ['MICROSOFT','MSFT34'],
    ['MICROSOFTDRN','MSFT34'],
    ['MILLSON','MILS3'],
    ['MINASMAQUINAON','MMAQ3'],
    ['MINASMAQUINAPN','MMAQ4'],
    ['MINASMAQUINASON','MMAQ3'],
    ['MINASMAQUINASPN','MMAQ4'],
    ['MINERVAON','BEEF3'],
    ['MINUPARON','MNPR3'],
    ['MITRE REALTYON','MTRE3'],
    ['MITREON','MTRE3'],
    ['Mitsubishi U DRN','M1UF34'],
    ['Mitsubishi U','M1UF34'],
    ['MMX MINERON','MMXM3'],
    ['MOBLYON','MBLY3'],
    ['Modalmais UNT N2','MODL11'],
    ['MODALMAISON','MODL3'],
    ['MODALMAISPN','MODL4'],
    ['MODERNA INC','M1RN34'],
    ['MODERNA INCDRN','M1RN34'],
    ['MONDELEZ INT','MDLZ34'],
    ['MONDELEZ INTDRN','MDLZ34'],
    ['Mongodb Inc DRN','M1DB34'],
    ['Mongodb Inc','M1DB34'],
    ['MONSTER BEVE','M1NS34'],
    ['MONSTER BEVEDRN','M1NS34'],
    ['MONT ARANHAON','MOAR3'],
    ['MORGAN STAN','MSBR34'],
    ['MORGAN STANDRN','MSBR34'],
    ['MOSAIC CO','MOSC34'],
    ['MOSAIC CODRN','MOSC34'],
    ['MOSAICO SAON','MOSI3'],
    ['MOSAICOON','MOSI3'],
    ['MOURA DUBEUXON','MDNE3'],
    ['MOVIDAON','MOVI3'],
    ['MRS LOGISTON','MRSA3B'],
    ['MRS LOGISTPNA','MRSA5B'],
    ['MRS LOGISTPNB','MRSA6B'],
    ['MRVON','MRVE3'],
    ['MULTILASERON','MLAS3'],
    ['MULTIPLANON','MULT3'],
    ['MUNDIALON','MNDL3'],
    ['NASDAQ INC','N1DA34'],
    ['NASDAQ INCDRN','N1DA34'],
    ['NEOENERGIAON','NEOE3'],
    ['NEOGRIDON','NGRD3'],
    ['NETEASE','NETE34'],
    ['NETEASEDRN','NETE34'],
    ['NETFLIX','NFLX34'],
    ['NETFLIXDRN','NFLX34'],
    ['NEW ORIENTAL','E1DU34'],
    ['NEW ORIENTALDRN','E1DU34'],
    ['NEWMONT GOLD','N1EM34'],
    ['NEWMONT GOLDDRN','N1EM34'],
    ['NEXTERA ENER','NEXT34'],
    ['NEXTERA ENERDRN','NEXT34'],
    ['NIKE','NIKE34'],
    ['NIKEDRN','NIKE34'],
    ['NOKIA CORP','NOKI34'],
    ['NOKIA CORPDRN','NOKI34'],
    ['NORD BRASILON','BNBR3'],
    ['NORDON METON','NORD3'],
    ['NORTCQUIMICAON','NRTQ3'],
    ['NORWEGIAN CR','N1CL34'],
    ['NORWEGIAN CRDRN','N1CL34'],
    ['NOVARTIS AG','N1VS34'],
    ['NOVARTIS AGDRN','N1VS34'],
    ['NOVO NORDISK','N1VO34'],
    ['NOVO NORDISKDRN','N1VO34'],
    ['NUTRIPLANTON','NUTR3'],
    ['NVIDIA CORP','NVDC34'],
    ['NVIDIA CORP','NVDC34'],
    ['NVIDIA CORPDRN','NVDC34'],
    ['Nxp Semicond DRN','N1XP34'],
    ['Nxp Semicond','N1XP34'],
    ['OCCIDENT PTR','OXYP34'],
    ['OCCIDENT PTRDRN','OXYP34'],
    ['OCEANPACTON','OPCT3'],
    ['ODERICHON','ODER3'],
    ['ODERICHPN','ODER4'],
    ['ODONTOPREVON','ODPV3'],
    ['OION','OIBR3'],
    ['OIPN','OIBR4'],
    ['Okta Inc DRN','O1KT34'],
    ['Okta Inc','O1KT34'],
    ['OMEGA GERON','OMGE3'],
    ['Omegaenergia ON NM','MEGA3'],
    ['ONCOCLINICASON','ONCO3'],
    ['OPPORT ENERGON','OPHE3B'],
    ['ORACLE','ORCL34'],
    ['ORACLEDRN','ORCL34'],
    ['ORIZONON','ORVR3'],
    ['OSX BRASILON','OSXB3'],
    ['OUROFINO S/AON','OFSA3'],
    ['P.ACUCAR-CBDON','PCAR3'],
    ['PACTUAL IBOV','IBOB11'],
    ['PACTUAL IBOVCI','IBOB11'],
    ['PADTECON','PDTC3'],
    ['Pagseguro DRN','PAGS34'],
    ['Pagseguro','PAGS34'],
    ['PAGUE MENOSON','PGMN3'],
    ['PANATLANTICAON','PATI3'],
    ['PANATLANTICAPN','PATI4'],
    ['PAR AL BAHIAON','PEAB3'],
    ['PAR AL BAHIAPN','PEAB4'],
    ['PARANAPANEMA','PMAM3'],
    ['PARANAPANEMAON','PMAM3'],
    ['PAYPAL HOLD','PYPL34'],
    ['PAYPAL HOLDDRN','PYPL34'],
    ['PBG S/AON','PTBL3'],
    ['PDG REALTON','PDGR3'],
    ['PEPSICO INC','PEPB34'],
    ['PEPSICO INCDRN','PEPB34'],
    ['PET MANGUINHON','RPMG3'],
    ['PETNGUINHON','RPMG3'],
    ['PETROBRAS BRON','BRDT3'],
    ['PETROBRASON','PETR3'],
    ['PETROBRASPN','PETR4'],
    ['PETROCHIN','PTCH34'],
    ['PETROCHINDRN','PTCH34'],
    ['PETRORECSAON','RECV3'],
    ['PETRORIOON','PRIO3'],
    ['PETTENATION','PTNT3'],
    ['PETTENATIPN','PTNT4'],
    ['PETZON','PETZ3'],
    ['PFIZER','PFIZ34'],
    ['PFIZERDRN','PFIZ34'],
    ['PG','PGCO34'],
    ['PGDRN','PGCO34'],
    ['PHILIP MORRI','PHMO34'],
    ['PHILIP MORRIDRN','PHMO34'],
    ['PINDUODUO IN','P1DD34'],
    ['PINDUODUO INDRN','P1DD34'],
    ['PINEON','PINE3'],
    ['PINEPN','PINE4'],
    ['PLANOEPLANOON','PLPL3'],
    ['PLASCAR PARTON','PLAS3'],
    ['POLPARON','PPAR3'],
    ['POMIFRUTASON','FRTA3'],
    ['PORTO SEGUROON','PSSA3'],
    ['PORTOBELLOON','PTBL3'],
    ['POSITIVO TECON','POSI3'],
    ['PPLAUNT','PPLA11'],
    ['PRATICAON','PTCA3'],
    ['PRATICAPN RESG','PTCA11'],
    ['PRINERON','PRNR3'],
    ['PROFARMAON','PFRM3'],
    ['PROLOGIS INC','P1LD34'],
    ['PROLOGIS INCDRN','P1LD34'],
    ['PROMPTON','PRPT3B'],
    ['Public Stora DRN','P1SA34'],
    ['Public Stora','P1SA34'],
    ['QR BITCOIN','QBTC11'],
    ['QR BITCOINCI','QBTC11'],
    ['QR DEFI','QDFI11'],
    ['QR DEFICI','QDFI11'],
    ['QR ETHER','QETH11'],
    ['QR ETHERCI','QETH11'],
    ['QUALCOMM','QCOM34'],
    ['QUALCOMMDRN','QCOM34'],
    ['QUALICORPON','QUAL3'],
    ['QUALITY SOFTON','QUSW3'],
    ['QUERO-QUEROON','LJQQ3'],
    ['RAIADROGASILON','RADL3'],
    ['RAIZENPN','RAIZ4'],
    ['RANDON PARTON','RAPT3'],
    ['RANDON PARTPN','RAPT4'],
    ['RAYTHEONTECH','RYTT34'],
    ['RAYTHEONTECHDRN','RYTT34'],
    ['RD SHELL','RDSA34'],
    ['RD SHELLDRN','RDSA34'],
    ['REALTY INCOM','R1IN34'],
    ['REALTY INCOMDRN','R1IN34'],
    ['RECRUSULON','RCSL3'],
    ['RECRUSULPN','RCSL4'],
    ['REDE D ORON','RDOR3'],
    ['REDE ENERGIAON','REDE3'],
    ['REGENERON PH','REGN34'],
    ['REGENERON PHDRN','REGN34'],
    ['RENOVAON','RNEW3'],
    ['RENOVAPN','RNEW4'],
    ['RENOVAUNT','RNEW11'],
    ['RIO TINTO','RIOT34'],
    ['RIO TINTODRN','RIOT34'],
    ['RIOSULENSEON','RSUL3'],
    ['RIOSULENSEPN','RSUL4'],
    ['RNION','RDNI3'],
    ['Rodobens UNT EJ','RBNS11'],
    ['ROKU INC','R1KU34'],
    ['ROKU INCDRN','R1KU34'],
    ['ROSSI RESIDON','RSID3'],
    ['ROYAL CARIBB','R1CL34'],
    ['ROYAL CARIBBDRN','R1CL34'],
    ['RUMO LOGON','RAIL3'],
    ['RUMO S.A.ON','RAIL3'],
    ['SABESPON','SBSP3'],
    ['SAFRAETFELAS','ELAS11'],
    ['SAFRAETFELASCI','ELAS11'],
    ['SAFRAETFIBOV','BOVS11'],
    ['SAFRAETFIBOVCI','BOVS11'],
    ['SALESFORCE','SSFO34'],
    ['SALESFORCEDRN','SSFO34'],
    ['SANEPARON','SAPR3'],
    ['SANEPARPN','SAPR4'],
    ['SANEPARUNT','SAPR11'],
    ['SANSUYON','SNSY3'],
    ['SANSUYPNA','SNSY5'],
    ['SANSUYPNB','SNSY6'],
    ['SANTANDER BRON','SANB3'],
    ['SANTANDER BRPN','SANB4'],
    ['SANTANDER BRUNT','SANB11'],
    ['SANTANDER','BCSA34'],
    ['SANTANDERDRN','BCSA34'],
    ['SANTANENSEON','CTSA3'],
    ['SANTANENSEPN','CTSA4'],
    ['SANTANENSEPND','CTSA8'],
    ['SANTOS BRPON','STBP3'],
    ['SAO CARLOSON','SCAR3'],
    ['SAO MARTINHOON','SMTO3'],
    ['SAORTINHOON','SMTO3'],
    ['SAP SE','SAPP34'],
    ['SAP SEDRN','SAPP34'],
    ['SARAIVA LIVRON','SLED3'],
    ['SARAIVA LIVRPN','SLED4'],
    ['SAUIPEON','PSEG3'],
    ['SAUIPEPN','PSEG4'],
    ['SCHULZON','SHUL3'],
    ['SCHULZPN','SHUL4'],
    ['SCHWAB','SCHW34'],
    ['SCHWABDRN','SCHW34'],
    ['Sea Ltd DRN','S2EA34'],
    ['Sea Ltd','S2EA34'],
    ['SEG AL BAHIAON','CSAB3'],
    ['SEG AL BAHIAPN','CSAB4'],
    ['SEQUOIA LOGON','SEQL3'],
    ['SER EDUCAON','SEER3'],
    ['SERUCAON','SEER3'],
    ['SERVICENOW','N1OW34'],
    ['SERVICENOWDRN','N1OW34'],
    ['Shopify Inc DRN','S2HO34'],
    ['Shopify Inc','S2HO34'],
    ['Sibanye Stil DRN','S1BS34'],
    ['Sibanye Stil','S1BS34'],
    ['SID NACIONALON','CSNA3'],
    ['SIMON PROP','SIMN34'],
    ['SIMON PROPDRN','SIMN34'],
    ['SIMPARON','SIMH3'],
    ['SINQIAON','SQIA3'],
    ['SL GREEN REA','S1LG34'],
    ['SL GREEN READRN','S1LG34'],
    ['SLC AGRICOLAON ED','SLCE3'],
    ['SLC AGRICOLAON','SLCE3'],
    ['SMART FITON','SMFT3'],
    ['SMART HASH','WEB311'],
    ['SMART HASHCI','WEB311'],
    ['SMILESON','SMLS3'],
    ['SONDOTECNICAON','SOND3'],
    ['SONDOTECNICAPNA','SOND5'],
    ['SONDOTECNICAPNB','SOND6'],
    ['SONY GROUP','SNEC34'],
    ['SONY GROUPDRN','SNEC34'],
    ['SPLUNK INC','S1PL34'],
    ['SPLUNK INCDRN','S1PL34'],
    ['SPOTIFY TECH','S1PO34'],
    ['SPOTIFY TECHDRN','S1PO34'],
    ['SPRINGSON','SGPS3'],
    ['SPTURISON','AHEB3'],
    ['SPTURISPNA','AHEB5'],
    ['SPTURISPNB','AHEB6'],
    ['STARAON','STTR3'],
    ['STARBUCKS','SBUB34'],
    ['STARBUCKSDRN','SBUB34'],
    ['STATKRAFTON','STKF3'],
    ['SUDESTEON','OPSE3B'],
    ['SUL 116 PARTON','OPTS3B'],
    ['SUL AMERICAON','SULA3'],
    ['SUL AMERICAPN','SULA4'],
    ['SUL AMERICAUNT','SULA11'],
    ['SUZANO HOLDON','NEMO3'],
    ['SUZANO HOLDPNA','NEMO5'],
    ['SUZANO HOLDPNB','NEMO6'],
    ['SUZANO PAPELON','SUZB3'],
    ['SUZANO PAPELPNA','SUZB5'],
    ['SUZANO S.A.ON','SUZB3'],
    ['SYN PROP TECON','SYNE3'],
    ['TAESAON','TAEE3'],
    ['TAESAPN','TAEE4'],
    ['TAESAUNT','TAEE11'],
    ['TAIWANSMFAC','TSMC34'],
    ['TAIWANSMFACDRN','TSMC34'],
    ['TAKE-TWO INT','T1TW34'],
    ['TAKE-TWO INTDRN','T1TW34'],
    ['TALUCATIO','T1AL34'],
    ['TALUCATIODRN','T1AL34'],
    ['TAPESTRY INC','TPRY34'],
    ['TAPESTRY INCDRN','TPRY34'],
    ['TARGET CORP','TGTB34'],
    ['TARGET CORPDRN','TGTB34'],
    ['TAURUS ARMASON','TASA3'],
    ['TAURUS ARMASPN','TASA4'],
    ['TCON','TRAD3'],
    ['TECH BRASIL','TECB11'],
    ['TECH BRASILCI','TECB11'],
    ['TECHNOSON','TECN3'],
    ['TECNISAON','TCSA3'],
    ['TECNOSOLOON','TCNO3'],
    ['TECNOSOLOPN','TCNO4'],
    ['TEGMA GESTAOON','TGMA3'],
    ['TEGMAON','TGMA3'],
    ['TEGRA INCORPON','TEGA3'],
    ['TEKAON','TEKA3'],
    ['TEKAPN','TEKA4'],
    ['TEKNOON','TKNO3'],
    ['TEKNOPN','TKNO4'],
    ['Teladochealt DRN','T2DH34'],
    ['Teladochealt','T2DH34'],
    ['TELEBRASON','TELB3'],
    ['TELEBRASPN','TELB4'],
    ['TELEF BRASILON','VIVT3'],
    ['TELEF BRASILPN','VIVT4'],
    ['TELEFONIC','TLNC34'],
    ['TELEFONICDRN','TLNC34'],
    ['TENDAON','TEND3'],
    ['Terniumsa DRN','TXSA34'],
    ['Terniumsa','TXSA34'],
    ['TERRA SANTA PA','LAND3'],
    ['TERRA SANTAON','TESA3'],
    ['TERRA SANTAPAON','LAND3'],
    ['TERRASANTAPA','LAND3'],
    ['TERRASANTAPAON','LAND3'],
    ['TESLA INC','TSLA34'],
    ['TESLA INCDRN','TSLA34'],
    ['Teva Pharmac DRN','T1EV34'],
    ['Teva Pharmac','T1EV34'],
    ['TEX RENAUXON','TXRX3'],
    ['TEX RENAUXPN','TXRX4'],
    ['TEXAS INC','TEXA34'],
    ['TEXAS INCDRN','TEXA34'],
    ['THERMFISCHER','TMOS34'],
    ['THERMFISCHERDRN','TMOS34'],
    ['TIM PARTON','TIMP3'],
    ['TIME FOR FUNON','SHOW3'],
    ['TIMON','TIMS3'],
    ['TOTVSON','TOTS3'],
    ['TOYOTAMO','TMCO34'],
    ['TOYOTAMODRN','TMCO34'],
    ['TRACK FIELDPN','TFCO4'],
    ['Trade Desk DRN','T2TD34'],
    ['Trade Desk','T2TD34'],
    ['TRAN PAULISTON','TRPL3'],
    ['TRAN PAULISTPN','TRPL4'],
    ['TRANSOCEAN','RIGG34'],
    ['TRANSOCEANDRN','RIGG34'],
    ['TREND ACWI','ACWI11'],
    ['TREND ACWICI','ACWI11'],
    ['TREND ASIA','ASIA11'],
    ['TREND ASIACI','ASIA11'],
    ['TREND CHINA','XINA11'],
    ['TREND CHINACI','XINA11'],
    ['TREND EMEG','EMEG11'],
    ['TREND EMEGCI','EMEG11'],
    ['TREND ESG D','ESGD11'],
    ['TREND ESG DCI','ESGD11'],
    ['TREND ESG E','ESGE11'],
    ['TREND ESG ECI','ESGE11'],
    ['TREND ESG US','ESGU11'],
    ['TREND ESG USCI','ESGU11'],
    ['TREND EUROPA','EURP11'],
    ['TREND EUROPACI','EURP11'],
    ['TREND IBOVX','BOVX11'],
    ['TREND IBOVXCI','BOVX11'],
    ['TREND IFIX-L','XFIX11'],
    ['TREND IFIX-LCI','XFIX11'],
    ['TREND NASDAQ','NASD11'],
    ['TREND NASDAQCI','NASD11'],
    ['TREND OURO','GOLD11'],
    ['TREND OUROCI','GOLD11'],
    ['TREND SMALL','XMAL11'],
    ['TREND SMALLCI','XMAL11'],
    ['TREND U REIT','URET11'],
    ['TREND U REITCI','URET11'],
    ['TREND US LRG','USAL11'],
    ['TREND US LRGCI','USAL11'],
    ['TREND US TEC','UTEC11'],
    ['TREND US TECCI','UTEC11'],
    ['TREVISAON','LUXM3'],
    ['TREVISAPN','LUXM4'],
    ['TRG SMIC CAP','TRIG11'],
    ['TRG SMIC CAPCI','TRIG11'],
    ['TRIPADVISOR','T1RI34'],
    ['TRIPADVISORDRN','T1RI34'],
    ['TRISUL S.A.ON','TRIS3'],
    ['TRISULON','TRIS3'],
    ['TRIUNFO PARTON','TPIS3'],
    ['TUPYON','TUPY3'],
    ['TWILIO INC','T1WL34'],
    ['TWILIO INCDRN','T1WL34'],
    ['TWITTER','TWTR34'],
    ['TWITTERDRN','TWTR34'],
    ['UBER TECH IN','U1BE34'],
    ['UBER TECH INDRN','U1BE34'],
    ['UBS GROUP','UBSG34'],
    ['UBS GROUPDRN','UBSG34'],
    ['ULTRAPARON','UGPA3'],
    ['UNIBANCOUNT','UBBR11'],
    ['UNICASAON','UCAS3'],
    ['UNIFIQUEON','FIQE3'],
    ['UNILEVER','ULEV34'],
    ['UNILEVERDRN','ULEV34'],
    ['UNIPAR CARBOPNB','UNIP6'],
    ['UNIPARON','UNIP3'],
    ['UNIPARPNA','UNIP5'],
    ['UNIPARPNB','UNIP6'],
    ['UNITED AIRLI','U1AL34'],
    ['UNITED AIRLIDRN','U1AL34'],
    ['UNITEDHEALTH','UNHH34'],
    ['UNITEDHEALTHDRN','UNHH34'],
    ['Unity Softwr DRN','U2ST34'],
    ['Unity Softwr','U2ST34'],
    ['UPS','UPSS34'],
    ['UPSDRN','UPSS34'],
    ['UPTICKON','UPKP3'],
    ['Us Bancorp DRN','USBC34'],
    ['Us Bancorp','USBC34'],
    ['US STEEL','USSX34'],
    ['US STEELDRN','USSX34'],
    ['USIMINASON','USIM3'],
    ['USIMINASPNA','USIM5'],
    ['USIMINASPNB','USIM6'],
    ['USINAS SID DPNA','USIM5'],
    ['VALE R DOCEPNA','VALE5'],
    ['VALEON','VALE3'],
    ['VALEPNA','VALE5'],
    ['VALERO ENER','VLOE34'],
    ['VALERO ENERDRN','VLOE34'],
    ['VALID SOLUÇÕON','VLID3'],
    ['VALIDON','VLID3'],
    ['VAMOSON','VAMO3'],
    ['VERIZON','VERZ34'],
    ['VERIZONDRN','VERZ34'],
    ['VERTEX PHARM','VRTX34'],
    ['VERTEX PHARMDRN','VRTX34'],
    ['VIACOMCBS','C1BS34'],
    ['VIACOMCBSDRN','C1BS34'],
    ['VIAON','VIIA3'],
    ['VIAVAREJOON','VVAR3'],
    ['VIBRAON','VBBR3'],
    ['Vipshop Hold DRN','V1IP34'],
    ['Vipshop Hold','V1IP34'],
    ['VISA INC','VISA34'],
    ['VISA INCDRN','VISA34'],
    ['VITTIAON','VITT3'],
    ['VIVARA S.A.ON','VIVA3'],
    ['VIVEOON','VVEO3'],
    ['VIVERON','VIVR3'],
    ['Vodafone Gro DRN','V1OD34'],
    ['Vodafone Gro','V1OD34'],
    ['VORNADO REAL','V1NO34'],
    ['VORNADO REALDRN','V1NO34'],
    ['VULCABRASON','VULC3'],
    ['WALGREENS','WGBA34'],
    ['WALGREENSDRN','WGBA34'],
    ['WALRT','WALM34'],
    ['WALRTDRN','WALM34'],
    ['WALT DISNEY','DISB34'],
    ['WALT DISNEYDRN','DISB34'],
    ['WDC NETWORKSON','LVTC3'],
    ['WEG S.A.ON','WEGE3'],
    ['WEGON ED','WEGE4'],
    ['WEGON','WEGE3'],
    ['Weibo Corp DRN','W1BO34'],
    ['Weibo Corp','W1BO34'],
    ['WELLS FARGO','WFCO34'],
    ['WELLS FARGODRN','WFCO34'],
    ['WESTWINGON','WEST3'],
    ['WETZEL S/AON','MWET3'],
    ['WETZEL S/APN','MWET4'],
    ['WHIRLPOOLON','WHRL3'],
    ['WHIRLPOOLPN','WHRL4'],
    ['WILSON SONSON','PORT3'],
    ['WILSON SONSON','WSON33'],
    ['WIX.COM LTD','W1IX34'],
    ['WIX.COM LTDDRN','W1IX34'],
    ['WIZ S.A.ON','WIZS3'],
    ['WIZ S.AON','WIZS3'],
    ['WLM IND COMON','WLMM3'],
    ['WLM IND COMPN','WLMM4'],
    ['YDUQS PARTON','YDUQ3'],
    ['ZOETIS INC','Z1TS34'],
    ['ZOETIS INCDRN','Z1TS34'],
    ['ZOOM VIDEO','Z1OM34'],
    ['ZOOM VIDEODRN','Z1OM34']
    ),
    columns=['TICKET','CODIGO'])

# ===================================================================================================
# Arquivo contendo o nome das principais Corretoras B3
# ATENÇÃO: Para NOVA corretora incluir o nome completo e o código no arquivo /Apoio/Corretoras.csv
# ===================================================================================================

#corretoras_cadastradas = pd.read_csv('./Apoio/Corretoras.csv')

# ===================================================================================================
# Há duas possibilidades de implementação. Buscando em um arquivo csv ou incorporando essas informações
# no próprio script python, conforme implementado a seguir. Analisar qual a melhor aboradagem.
# ===================================================================================================
''' Atualizado em 31/01/2022'''
corretoras_cadastradas = pd.DataFrame(data=(
    ['3','XP INVESTIMENTOS CCTVM S/A','XP'],
    ['386','RICO INVESTIMENTOS - GRUPO XP','RICO'],
    ['90','EASYNVEST - TITULO CV S.A.','EASYNVEST'],
    ['308','CLEAR CORRETORA - GRUPO XP','CLEAR'],
    ['85','BTG PACTUAL CTVM S.A.','BTG'],
    ['72','BRADESCO S/A CTVM','BRADESCO'],
    ['39','AGORA CTVM S/A','AGORA'],
    ['39','AGORA CORRETORA DE TITULOS E VALORES MOBILIARIOS S/A','AGORA'],
    ['3701','ORAMA DTVM S.A.','ORAMA'],
    ['735','ICAP DO BRASIL CTVM LTDA','ICAP'],
    ['120','GENIAL INSTITUCIONAL CCTVM S.A','GENIAL'],
    ['173','GENIAL INVESTIMENTOS CVM S.A.','GENIAL'],
    ['93','NOVA FUTURA CTVM LTDA','NOVA FUTURA'],
    ['107','TERRA INVESTIMENTOS DTVM LTDA','TERRA'],
    ['6003','C6 CTVM LTDA','C6'],
    ['1982','MODAL DTVM LTDA','MODAL'],
    ['683','BANCO MODAL','MODAL'],
    ['4090','TORO CTVM LTDA.','TORO'],
    ['1099','INTER DTVM LTDA','INTER'],
    ['15','GUIDE INVESTIMENTOS S.A. CV','GUIDE'],
    ['114','ITAU CV S/A','ITAU'],
    ['820','BB BANCO DE INVESTIMENTO S/A','BB'],
    ['713','BB GESTAO DE RECURSOS DTVM S/A','BB'],
    ['147','ATIVA INVESTIMENTOS S.A. CTCV','ATIVA'],
    ['1618','IDEAL CTVM SA','IDEAL'],
    ['172','BANRISUL S/A CVMC','BANRISUL'],
    ['442','BANCO OURINVEST','OURINVEST'],
    ['359','BANCO DAYCOVAL','DAYCOVAL'],
    ['1116','BANCO CITIBANK','CITIBANK'],
    ['251','BANCO BNP PARIBAS BRASIL S/A','BNP'],
    ['4','ALFA CCVM S.A.','ALFA'],
    ['979','ADVALOR DTVM LTDA','ADVALOR'],
    ['226','AMARIL FRANKLIN CTV LTDA.','AMARIL'],
    ['4002','BANCO ANDBANK (BRASIL) S.A.','ANDBANK'],
    ['3112','BANESTES DTVM S/A','BANESTES'],
    ['2197','BCO FIBRA','FIBRA'],
    ['122','BGC LIQUIDEZ DTVM','BGC'],
    ['18','BOCOM BBM CCVM S/A','BOCOM'],
    ['4015','BS2 DTVM S/A','BS2'],
    ['1570','CAIXA ECONOMICA FEDERAL','CEF'],
    ['77','CITIGROUP GMB CCTVM S.A.','CITIGROUP'],
    ['88','CM CAPITAL MARKETS CCTVM LTDA','CM CAPITAL'],
    ['234','CODEPE CV E CAMBIO S/A','CODEPE'],
    ['74','COINVALORES CCVM LTDA.','COINVALORES'],
    ['186','CORRETORA GERAL DE VC LTDA','GERAL'],
    ['45','CREDIT SUISSE BRASIL S.A. CTVM','CREDIT SUISSE'],
    ['133','DIBRAN DTVM LTDA','DIBRAN'],
    ['711','DILLON S.A. DTVM','DILLON'],
    ['174','ELITE CCVM LTDA.','ELITE'],
    ['131','FATOR S.A. CV','FATOR'],
    ['238','GOLDMAN SACHS DO BRASIL CTVM','GOLDMAN'],
    ['115','H.COMMCOR DTVM LTDA','H.COMMCOR'],
    ['41','ING CCT S/A','ING'],
    ['1130','INTL FCSTONE DTVM LTDA.','INTL'],
    ['16','J. P. MORGAN CCVM S.A.','J. P. MORGAN'],
    ['33','LEROSA S.A. CVC','LEROSA'],
    ['2640','LLA DTVM LTDA','LLA'],
    ['1','MAGLIANO S.A. CCVM','MAGLIANO'],
    ['83','MAXIMA S/A CTVM','MAXIMA'],
    ['106','MERC. DO BRASIL COR. S.A. CTVM','MERC.'],
    ['13','MERRILL LYNCH S/A CTVM','MERRILL'],
    ['262','MIRAE ASSET WEALTH MANAGEMENT','MIRAE'],
    ['40','MORGAN STANLEY CTVM S/A','MORGAN'],
    ['181','MUNDINVEST S.A. CCVM','MUNDINVEST'],
    ['23','NECTON INVESTIMENTOS S.A. CVMC','NECTON'],
    ['63','NOVINVEST CVM LTDA.','NOVINVEST'],
    ['2379','ORLA DTVM S/A','ORLA'],
    ['1106','OURINVEST DTVM S.A.','Ourinvest'],
    ['129','PLANNER CV S.A','PLANNER'],
    ['2492','POSITIVA CTVM S/A','POSITIVA'],
    ['1089','RB CAPITAL INVESTIMENTOS DTVM','RB CAPITAL'],
    ['92','RENASCENCA DTVM LTDA.','RENASCENCA'],
    ['3371','RIO BRAVO INVEST S.A. DTVM','RIO BRAVO'],
    ['3762','RJI CTVM LTDA','RJI'],
    ['59','SAFRA CVC LTDA.','SAFRA'],
    ['27','SANTANDER CCVM S/A','SANTANDER'],
    ['2570','SANTANDER SECURITIES SERVICES','SANTANDER'],
    ['191','SENSO CCVM S.A.','SENSO'],
    ['187','SITA SCCVM S.A.','SITA'],
    ['110','SLW CVC LTDA.','SLW'],
    ['58','SOCOPA SC PAULISTA S.A.','SOCOPA'],
    ['177','SOLIDUS S/A CCVM','SOLIDUS'],
    ['127','TULLETT PREBON','TULLETT'],
    ['8','UBS BRASIL CCTVM S/A','UBS'],
    ['37','UM INVESTIMENTOS S.A. CTVM','UM INVESTIMENTOS'],
    ['29','UNILETRA CCTVM S.A.','UNILETRA'],
    ['21','VOTORANTIM ASSET MANAG. DTVM','VOTORANTIM']
    ),
    columns=['Codigo','Corretora','Nome'])

# ===================================================================================================
# Padrão de leitura dos arquivos PDF's contendo as Notas de Corretagem no padrão SINANCOR
# ===================================================================================================
#col1str = {'dtype': str}
col1str = {'header': None}
kwargs = {
        'multiple_tables':False,
        'encoding': 'utf-8',
        'pandas_options': col1str,
        'stream':True,
        'guess':False
}

#rxcountpages = re.compile(r"/Type\s*/Page([^s]|$)", re.MULTILINE|re.DOTALL)
#def count_pages(filename):
#    data = file(filename,"rb").read()
#    return len(rxcountpages.findall(data))
#

# ===================================================================================================
# Formata valores floats para uma quantidade específica de casas decimais
# ===================================================================================================
def truncate(number, decimals=0):
    from math import trunc
    if decimals < 0:
        raise ValueError('truncate received an invalid value of decimals ({})'.format(decimals))
    elif decimals == 0:
        return trunc(number)
    else:
        factor = float(10**decimals)
        return trunc(number*factor)/factor

'''Os valores de PM e total da SENCON não tem limite de casas decimais.
Após teste de controle foi observado que TRUNCAR o PM na planilha COIR ficava com alguns centavos 
a menor ou maior em relação a SENCON (sencon.com.br)'''
        
# ===================================================================================================
# Formata campos contendo strings para float
# ===================================================================================================
def sanitiza_moeda(moeda):
    #Se o valor é uma string então remove os simbolo $ e delimitadores
    #caso contrário, o valor é numérico e pode ser convertido
    if isinstance(moeda, str):
        return moeda.replace('CONTINUA...','0').replace('T - Liquidação pelo Bruto','0').replace('ON NM','0').replace('.','').replace(',','.').replace('R$','').replace('$','').replace('NM','0').replace('ON','0').replace('N1','0').replace('N2','0').replace('C O N T I N U A   ','0').replace("| D",'').replace("|D ",'').replace(" |D",'').replace('D','').replace("| C",'').replace("|C",'').replace("|",'').replace("| ",'').replace("0| ",'0').replace("0|",'0').replace('Compra Opções','0').replace('0 ay Trade (proj)','0').replace('+0 Custos Impostos','0')
    return moeda
    #return moeda.replace('CONTINUA...','0').replace('T - Liquidação pelo Bruto','0').replace('ON NM','0').replace('.','').replace(',','.').replace('R$','').replace('$','').replace('NM','0').replace('ON','0').replace('N1','0').replace('N2','0').replace('C O N T I N U A   ','0')
    
# ===================================================================================================
# Formata campos das Notas BMF com mais de uma página por nota
# ===================================================================================================
def sanitiza_nota_bmf(value):
    if isinstance(value, str):
        return value.replace('Venda disponível','0').replace('IRRF','0').replace('Outros','0').replace('Compra disponível','0').replace('IRRF Day Trade (proj.)','0').replace('0 ay Trade (proj)','0').replace('+Outros Custos Impostos','0').replace('Total Conta Investimento','0').replace('Venda Opções','0').replace('Taxa operacional','0').replace('Ajuste de posição','0').replace('Total Conta Normal','0').replace('Compra Opções','0').replace('Taxa registro BM&F','0').replace('Ajuste day trade','0').replace('Total liquido (#)','0').replace('Valor dos negócios','0').replace('Taxas BM&F (emol+f.gar)','0').replace('Total das despesas','0').replace('Total líquido da nota','0')
    return value

# ===================================================================================================
# Sanitinizar Especificação do título
# ===================================================================================================
def sanitiza_especificacao_titulo(especificacao_titulo):
    especificacao_titulo = especificacao_titulo.str.replace(' NM','',regex=False).str.replace('  N1','',regex=False).str.replace(' N1','',regex=False).str.replace('N1','',regex=False).str.replace(' N2','',regex=False).str.replace('  N2','',regex=False).str.replace(' EDJ','',regex=False).str.replace(' EDB','',regex=False).str.replace(' ED','',regex=False).str.replace(' MA','',regex=False).str.replace(' M2 ','',regex=False).str.replace(' M2','',regex=False).str.replace(' MB','',regex=False).str.replace(' DR1','',regex=False).str.replace(' DR2','',regex=False).str.replace(' DR3','',regex=False).str.replace(' DRE','',regex=False).str.replace('DRN A','',regex=False).str.replace('DRN','',regex=False).str.replace(' EJS','',regex=False).str.replace(' EJ','',regex=False).str.replace(' EB','',regex=False).str.replace(' *','',regex=False)
    return especificacao_titulo
    #especificacao_titulo = especificacao_titulo.str.replace(' NM','',regex=False).str.replace('  N1','',regex=False).str.replace(' N1','',regex=False).str.replace('N1','',regex=False).str.replace(' N2 ','',regex=False).str.replace(' N2','',regex=False).str.replace(' EDJ','',regex=False).str.replace(' EDB','',regex=False).str.replace(' ED','',regex=False).str.replace(' MA','',regex=False).str.replace(' M2 ','',regex=False).str.replace(' M2','',regex=False).str.replace(' MB','',regex=False).str.replace(' DR1','',regex=False).str.replace(' DR2','',regex=False).str.replace(' DR3','',regex=False).str.replace(' DRE','',regex=False).str.replace('DRN A','',regex=False).str.replace('DRN','',regex=False).str.replace(' EJS','',regex=False).str.replace(' EJ','',regex=False).str.replace(' EB','',regex=False).str.replace(' *','',regex=False)
    
# ===================================================================================================
# Sanitinizar observação
# ===================================================================================================
def sanitiza_observacao(observacao):
    observacao = observacao.str.replace('A','').str.replace('T','').str.replace('#2','').str.replace('2','').str.replace('C','').str.replace('I','').str.replace('#','').str.replace('P','').str.replace('8','').str.replace('H','').str.replace('X','').str.replace('F','').str.replace('Y','').str.replace('B','').str.replace('L','')
    return observacao

# ===================================================================================================
# Quantidade operada de cada ativo por nota de corretagem
# ===================================================================================================
def quantidade_operada(df_quantidade=0,df_unnamed_0=0,df_unnamed_1=0,df_unnamed_2=0):
    quantidade = 0
    if df_quantidade > 0:
        quantidade = df_quantidade
    elif isinstance(df_unnamed_0, float) and df_unnamed_0 > 0:
        quantidade = df_unnamed_0
    elif df_unnamed_1 > 0:
        quantidade = df_unnamed_1
    elif df_unnamed_2 > 0:
        quantidade = df_unnamed_2
    return quantidade
    #try:
    #    if df_unnamed_0 > 0:
    #        df_unnamed_0 = df_unnamed_0
    #except:
    #    df_unnamed_0 = 0

# ===================================================================================================
# Valor total de cada operação por nota de corretagem
# ===================================================================================================
def valor_total_ativo(df_valor_operacao_ajuste,df_unnamed_2=0,df_unnamed_1=0):
    total = 0
    if df_valor_operacao_ajuste > 0:
        total = df_valor_operacao_ajuste
    elif df_unnamed_2 > 0:
        total = df_unnamed_2
    elif df_unnamed_1 > 0:
        total = df_unnamed_1
    return total

# ===================================================================================================
# Obter o preço unitário de cada operação
# ===================================================================================================
'''Foi excluída a etapa para extração do preço unitário de cada operação.
   Pois esse dado é obtido pela divisão do valor_total / quantidade.'''
#def obter_preco():
#    if df['Preço / Ajuste'].iloc[current_row] > 0:
#        price = df['Preço / Ajuste'].iloc[current_row]
#    elif 'Unnamed: 0' in df.columns:
#        if df['Unnamed: 0'].iloc[current_row] > 0 and df['Quantidade'].iloc[current_row] > 0:
#            price = df['Unnamed: 0'].iloc[current_row]
#        elif 'Unnamed: 1' in df.columns:
#            if df['Unnamed: 1'].iloc[current_row] > 0:
#                price = df['Unnamed: 1'].iloc[current_row]
#            elif 'Unnamed: 2' in df.columns:
#                if df['Unnamed: 2'].iloc[current_row] > 0:
#                    price = df['Unnamed: 2'].iloc[current_row]

# ===================================================================================================
# Dividindo os custos operacionais e o IRRF por cada operação.
# ===================================================================================================
def custos_por_operacao(taxas_df,number,c_v,total,operacao):
    custos_fin = 0
    irrf = 0
    for current_row in taxas_df.index:
        cell_value = taxas_df['Nota'].iloc[current_row]
        if cell_value == number:
            custos_fin = (total/taxas_df['Total'].iloc[current_row])*taxas_df['Custos_Fin'].iloc[current_row]
            if c_v == 'V' and operacao == "Normal":
                if taxas_df['BaseCalculo'].iloc[current_row] > 0:
                    irrf = (total/taxas_df['BaseCalculo'].iloc[current_row])*taxas_df['IRRF'].iloc[current_row]
                else:
                    irrf = (total/taxas_df['Vendas'].iloc[current_row])*taxas_df['IRRF'].iloc[current_row]
            else:
                irrf = 0
    return custos_fin,irrf

# ===================================================================================================
# Dividindo os custos operacionais e o IRRF por cada operação BMF e Contratos Futuros.
# ================================================================================================
def custos_por_operacao_bmf(taxas_df,number,corretagem):
    custos_fin = 0
    for current_row in taxas_df.index:
        cell_value = taxas_df['Nota'].iloc[current_row]
        if cell_value == number:
            custos_fin = (corretagem/taxas_df['Corretagem'].iloc[current_row])*taxas_df['Custos_Fin'].iloc[current_row]
    return custos_fin
    
# ===================================================================================================
# Altera o nome de papel conforme as atualizações das empresas na B3
# ===================================================================================================
def altera_ticket(stock, data):
    log = ''
    if stock == "JSLG3":
        datalimite = datetime.strptime("19/09/2020", '%d/%m/%Y').date()
        if data < datalimite:
            stock = "SIMH3"
            log = 'As operações anteriores ao dia 19/09/2020 com o código "JSLG3" foram alteradas para o código "SIMH3".\n'
    return stock,log

# ===================================================================================================
# Altera o nome do papel no pregão para o código no padrão XXXX4, por exemplo "PETR4", "VALE3", etc.
# ===================================================================================================
def nome_pregao(acoes, stock_title, data):
    control = 0
    log = ''
    for current_row in acoes.index:
        cell_value = acoes['TICKET'].iloc[current_row]        
        if cell_value == stock_title:
            stock_title = acoes['CODIGO'].iloc[current_row]
            control = 1
            break
    stock_title,log_altera_ticket = altera_ticket(stock_title, data)
    log += log_altera_ticket
    
    #Mensagem de alerta da variável stock_title, caso não tenha sido encontrada, conforem laço anterior
    if control == 0:
        print_atencao()
        #print('\n')
        print('O ativo',stock_title,'ainda não foi cadastrado!"')
        log += 'ATENÇÃO:\n'
        log += ' - O ativo '+stock_title+' ainda não foi cadastrado!\n'
        stock_title = stock_title[0:5] + '99'
        print('Na planilha Normais_Dados ou DayTrade_Dados do arquivo COIR.xlsb irá aparecer o código',stock_title)
        print('Realize a alteração para o código correto do ativo APENAS nessas planilhas.','\n')
        log += ' - Na planilha Normais_Dados ou DayTrade_Dados do arquivo COIR.xlsb irá aparecer o código: "'+stock_title+'".\n'
        log += ' - Realize a alteração para o código correto do ativo APENAS nessas planilhas.\n'
    return stock_title,log

# ===================================================================================================
# Obtem o valor total de cada operação, o multipicador e o código B3.
# ===================================================================================================    
def mercadoria_ticket(mercadoria, preco_unitario, quantidade):        
    if 'CCM' in mercadoria:
        valor_total = preco_unitario * quantidade * 450
        id = 'CCM'
        mult = 450
    elif 'BGI' in mercadoria:
        valor_total = preco_unitario * quantidade * 330
        id = 'BGI'
        mult = 330
    elif 'ICF' in mercadoria: #cotação variável conforme o dolar
        valor_total = preco_unitario * quantidade * 100 # * dolar
        id = 'ICF'
        mult = 100
        print_atencao()
        print('A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!"')
        print('É importante verificar nas planilhas de operações se os valores estão corretos.')
        log += 'ATENÇÃO:\n'
        log += ' - A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!\n'
        log += ' - É importante verificar nas planilhas de operações se os valores estão corretos.\n'
    elif 'SJC' in mercadoria: #cotação variável conforme o dolar
        valor_total = preco_unitario * quantidade * 450 # * dolar
        id = 'SJC'
        mult = 450
        print_atencao()
        print('A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!"')
        print('É importante verificar nas planilhas de operações se os valores estão corretos.')
        log += 'ATENÇÃO:\n'
        log += ' - A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!\n'
        log += ' - É importante verificar nas planilhas de operações se os valores estão corretos.\n'
    elif 'WIN' in mercadoria:
        valor_total = preco_unitario * quantidade * 0.2
        id = 'WIN'
        mult = 0.2
    elif 'IND' in mercadoria:
        valor_total = preco_unitario * quantidade * 1
        id = 'IND'
        mult = 1
    elif 'WDO' in mercadoria:
        valor_total = preco_unitario * quantidade * 10
        id = 'WDO'
        mult = 10
    elif 'DOL' in mercadoria:
        valor_total = preco_unitario * quantidade * 50
        id = 'DOL'
        mult = 50
    return valor_total,id,mult

# ===================================================================================================
#Contabiliza a quantidade de vendas e o valor de IR nas operações DayTrade e Normal BM&F
# ===================================================================================================  
def ir_bmf(cont_notas,note_df,taxas_df,row_data,note_data):
    cont_note = len(note_df['Nota'])
    #print(note_df.head(40))
    for t in range(0,cont_notas):
        cont_v_daytrade = 0
        cont_v_normal = 0
        cont_c_normal = 0
        if taxas_df['IRRF'].iloc[t] > 0:
            for n in range(0,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'V':
                    cont_v_normal += 1
            if cont_v_normal == 0:
                for n in range(0,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'C':
                        cont_c_normal += 1
            if cont_v_normal > 0:
                for n in range(0,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'V':
                        #note_df['IRRF'].iloc[n] = taxas_df['IRRF'].iloc[t] / cont_v_normal
                        row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IRRF'].iloc[t]/cont_v_normal, 0,0,0]
                        note_data.append(row_data)
            elif cont_c_normal > 0:
                for n in range(0,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'C':
                        #note_df['IRRF'].iloc[n] = taxas_df['IRRF'].iloc[t] / cont_c_normal
                        row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IRRF'].iloc[t]/cont_c_normal, 0,0,0]
                        note_data.append(row_data)
        elif taxas_df['IR_DT'].iloc[t] > 0:
            for n in range(0,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'DayTrade' and note_df['C/V'].iloc[n] == 'V':
                    cont_v_daytrade += 1
            for n in range(0,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'DayTrade' and note_df['C/V'].iloc[n] == 'V':
                    #note_df['IRRF'].iloc[n] = taxas_df['IR_DT'].iloc[t] / cont_v_daytrade
                    '''Testar com essa nova substituição de campo'''
                    #note_df['IR_DT'].iloc[n] = taxas_df['IR_DT'].iloc[t] / cont_v_daytrade
                    row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IR_DT'].iloc[t]/cont_v_daytrade, 0,0,0]
                    note_data.append(row_data)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço','Quantidade', 'Total','Custos_Fin','PM','IRRF','IR_DT','ID','FATOR']
    note_df = pd.DataFrame(data=note_data, columns=cols)  
    '''
    for t in range(0,cont_notas):
        cont_v_daytrade = 0
        cont_v_normal = 0
        cont_c_normal = 0
        if taxas_df['IRRF'].iloc[t] > 0:
            for n in range(1,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'V':
                    cont_v_normal += 1
            if cont_v_normal == 0:
                for n in range(1,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'C':
                        cont_c_normal += 1
            if cont_v_normal > 0:
                for n in range(1,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'V':
                        #note_df['IRRF'].iloc[n] = taxas_df['IRRF'].iloc[t] / cont_v_normal
                        row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IRRF'].iloc[t]/cont_v_normal, 0,0,0]
                        note_data.append(row_data)
            elif cont_c_normal > 0:
                for n in range(1,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'C':
                        #note_df['IRRF'].iloc[n] = taxas_df['IRRF'].iloc[t] / cont_c_normal
                        row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IRRF'].iloc[t]/cont_c_normal, 0,0,0]
                        note_data.append(row_data)
        elif taxas_df['IR_DT'].iloc[t] > 0:
            for n in range(1,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'DayTrade' and note_df['C/V'].iloc[n] == 'V':
                    cont_v_daytrade += 1
            for n in range(1,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'DayTrade' and note_df['C/V'].iloc[n] == 'V':
                    #note_df['IRRF'].iloc[n] = taxas_df['IR_DT'].iloc[t] / cont_v_daytrade
                    Testar com essa nova substituição de campo
                    #note_df['IR_DT'].iloc[n] = taxas_df['IR_DT'].iloc[t] / cont_v_daytrade
                    row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IR_DT'].iloc[t]/cont_v_daytrade, 0,0,0]
                    note_data.append(row_data)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço','Quantidade', 'Total','Custos_Fin','PM','IRRF','IR_DT','ID','FATOR']
    note_df = pd.DataFrame(data=note_data, columns=cols)
    #print(note_df.head(40))
    '''
    #print(note_df.head(40))
    return(note_df)

# ===================================================================================================
# Obtendo o valor preço médio de cada operação
# ===================================================================================================
def preco_medio(c_v,total,custos_fin,quantidade):
    preco_medio = 0
    if c_v == "C":
        preco_medio = round((total + custos_fin)/quantidade,4)
    elif c_v =="V":
        preco_medio = round((total - custos_fin)/quantidade,4)
    return preco_medio
    
# ===================================================================================================
# Validação de notas de corretagem no padrão Sinacor
# ===================================================================================================
#def valida_nota_corretagem(validacao,item): 
#    try:
#        df_validacao = pd.concat(validacao,axis=1,ignore_index=True)
#        df_validacao = pd.DataFrame({'NotaCorretagem': df_validacao[0].unique()})
#           
#                                                       
#        cell_value = df_validacao['NotaCorretagem'].iloc[0]
#                                                                    
#        if cell_value == 'NOTA DE NEGOCIAÇÃO' or 'NOTA DE CORRETAGEM':
#            print('processando o arquivo:',basename(item))
#        else:
#            print_atencao()
#            print('O arquivo','"'+basename(item).upper()+'"','NÃO é uma Nota de Corretagem no Padrão Sinacor.','\n')
#            return continue
#    except ValueError:
#        print_atencao()
#        print('O arquivo','"'+basename(item).upper()+'"','NÃO é uma Nota de Corretagem no Padrão Sinacor.','\n')
#        return continue

def valida_nota_corretagem(validacao,item): 
    df_validacao = pd.concat(validacao,axis=1,ignore_index=True)
    df_validacao = pd.DataFrame({'NotaCorretagem': df_validacao[0].unique()})
    cell_value = df_validacao['NotaCorretagem'].iloc[0]
    if cell_value == 'NOTA DE NEGOCIAÇÃO' or 'NOTA DE CORRETAGEM':
        print('processando o arquivo:',basename(item))
    #return()
    '''Função não está em uso
    os comando de continue não pode ser utilizados fora de laços'''

# ===================================================================================================
# Validação de notas de corretagem no padrão Sinacor
# ===================================================================================================
def valida_corretora(corretora):
    control = 0
    df_corretora = pd.concat(corretora,axis=1,ignore_index=True)
    cell_value = df_corretora[1].iloc[3]
    cell_value = cell_value.upper()    
    if type(df_corretora[0].iloc[3]) == str:
        nota_BMF = df_corretora[0].iloc[3]
        nota_BMF = nota_BMF.upper()       
    for current_row in corretoras_cadastradas.index:
        corretora_value = corretoras_cadastradas['Corretora'].iloc[current_row]
        if cell_value == corretora_value:
            corretora = corretoras_cadastradas['Nome'].iloc[current_row]
            control = 1
            break
    if control == 0:
        for current_row in corretoras_cadastradas.index:
            corretora_value = corretoras_cadastradas['Corretora'].iloc[current_row]
            if nota_BMF == corretora_value:
                corretora = corretoras_cadastradas['Nome'].iloc[current_row]
                control = 2
                break    
    return control,corretora,cell_value

# ===================================================================================================
# Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
# ===================================================================================================
def agrupar(note_df):
    note_df_agrupado = note_df.groupby(
    ['Corretora','CPF','Nota', 'Data', 'C/V', 'Papel', 'Operacao'],as_index=False
    ).agg(
        {
            'Preço': sum,
            'Quantidade': sum,
            'Total': sum,
            'Custos_Fin': sum,
            'PM': sum,
            'IRRF': sum
        }
    )
    return note_df_agrupado

# ===================================================================================================
# Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
# ===================================================================================================
def agrupar_bmf(note_df):
    note_df_agrupado = note_df.groupby(
    ['Corretora','CPF','Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço'],as_index=False
    ).agg(
        {
            'Quantidade': sum,
            'Total': sum,
            'Custos_Fin': sum,
            'PM': sum,
            'IRRF': sum,
            'IR_DT': sum
        }
    )
    return note_df_agrupado

# ===================================================================================================
# Seleção de papel isento de IR (IRRF e IRPF)
# Caso haja mais de um papel isento em uma mesma NC o sistema NÃO detectará
# A implementação de analisar mais de um papel isento por NC continua pendente
# ===================================================================================================     
def isencao_imposto_renda(taxas_df,grouped,note_data):
    log = ''
    controle = 0
    #Analisa as operações com uma venda isenta de IR
    for current_row in taxas_df.index:
        if (taxas_df['Vendas'].iloc[current_row] and taxas_df['BaseCalculo'].iloc[current_row]) > 0 and taxas_df['Vendas'].iloc[current_row] != taxas_df['BaseCalculo'].iloc[current_row]:
            diff = taxas_df['Vendas'].iloc[current_row] - taxas_df['BaseCalculo'].iloc[current_row]
            for I in range(0,len(grouped)):
                if grouped['Total'].iloc[I] == diff and grouped['Operacao'].iloc[I] == 'Normal':
                    note_data.append([grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'V',grouped['Papel'].iloc[I],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[I])*-1])
                    datetime.today().strftime('%d/%m/%Y')
                    log = 'Na operação de venda de "'+str(grouped["Papel"].iloc[I])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[I]))+', do dia '+str(grouped["Data"].iloc[I].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                    controle = 1   
            #Analisa as operações com duas vendas isentas de IR
            if controle == 0:
                for i in range(0,len(grouped)):
                    for p in range(grouped.index[i],len(grouped)):
                        if grouped['Total'].iloc[i] + grouped['Total'].iloc[p] == diff and grouped['Operacao'].iloc[i] == 'Normal':
                            note_data.append([grouped['Corretora'].iloc[i],grouped['CPF'].iloc[i],grouped['Nota'].iloc[i],grouped['Data'].iloc[i],'V',grouped['Papel'].iloc[i],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[i])*-1])
                            note_data.append([grouped['Corretora'].iloc[p],grouped['CPF'].iloc[p],grouped['Nota'].iloc[p],grouped['Data'].iloc[p],'V',grouped['Papel'].iloc[p],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[p])*-1])
                            datetime.today().strftime('%d/%m/%Y')
                            log = 'Na operação de venda de "'+str(grouped["Papel"].iloc[i])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[i]))+', do dia '+str(grouped["Data"].iloc[i].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                            log += 'Na operação de venda de "'+str(grouped["Papel"].iloc[p])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[p]))+', do dia '+str(grouped["Data"].iloc[p].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                            controle = 1
                            break
                    else:
                        continue
                    break
            #Analisa as operações com três vendas isentas de IR
            if controle == 0:
                for i in range(0,len(grouped)):
                    for p in range(grouped.index[i],len(grouped)):
                        for z in range(grouped.index[p],len(grouped)):
                            if grouped['Total'].iloc[i] + grouped['Total'].iloc[p] + grouped['Total'].iloc[z] == diff and grouped['Operacao'].iloc[i] == 'Normal':
                                note_data.append([grouped['Corretora'].iloc[i],grouped['CPF'].iloc[i],grouped['Nota'].iloc[i],grouped['Data'].iloc[i],'V',grouped['Papel'].iloc[i],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[i])*-1])
                                note_data.append([grouped['Corretora'].iloc[p],grouped['CPF'].iloc[p],grouped['Nota'].iloc[p],grouped['Data'].iloc[p],'V',grouped['Papel'].iloc[p],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[p])*-1])
                                note_data.append([grouped['Corretora'].iloc[z],grouped['CPF'].iloc[z],grouped['Nota'].iloc[z],grouped['Data'].iloc[z],'V',grouped['Papel'].iloc[z],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[z])*-1])
                                datetime.today().strftime('%d/%m/%Y')
                                log = 'Na operação de venda de "'+str(grouped["Papel"].iloc[i])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[i]))+', do dia '+str(grouped["Data"].iloc[i].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                                log += 'Na operação de venda de "'+str(grouped["Papel"].iloc[p])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[p]))+', do dia '+str(grouped["Data"].iloc[p].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                                log += 'Na operação de venda de "'+str(grouped["Papel"].iloc[z])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[z]))+', do dia '+str(grouped["Data"].iloc[z].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                                controle = 1                               
                                break
                        else:
                            continue
                        break
                    else:
                        continue
                    break
    return note_data,log

# ===================================================================================================
# Agrupa as operações por tipo de operação (Normal ou Daytrade)
# ===================================================================================================
def agrupar_operacoes(grouped,cols):
    groups = grouped.groupby(grouped.Operacao)
    #result = 0
    try:
        normal_df = groups.get_group("Normal")
    except KeyError:
        normal_df = pd.DataFrame(columns=cols)
    try:
        daytrade_df = groups.get_group("DayTrade")
        vendas = daytrade_df.loc[daytrade_df['C/V'] == 'V']
        compras = daytrade_df.loc[daytrade_df['C/V'] == 'C']
        result = pd.merge(compras, vendas, on=["Corretora", "CPF","Nota","Data","Papel",'Operacao','IRRF'])
        result['QTDE'] = result['Quantidade_x'] - result['Quantidade_y']
        result['Lucro'] = (result['Total_y']/result['Quantidade_y'] - result['Total_x']/result['Quantidade_x'])*((result['QTDE']+result['Quantidade_x']+result['Quantidade_y'])/2)
        return normal_df,daytrade_df,result
    except KeyError:
        daytrade_df = pd.DataFrame(columns=cols)
        return normal_df,daytrade_df
        
    #if 'result' in locals():
    #    return normal_df,daytrade_df,result
    #else:
    #    return normal_df,daytrade_df

# ===================================================================================================
# Insere o valor do IR para as operações de Daytrade"
# ===================================================================================================
def daytrade_ir(result,taxas_df,note_data,grouped):
    log = ''
    result = result
    for current_row in result.index:
        if result['Lucro'].iloc[current_row] > 0:
            for I in taxas_df.index:
                if taxas_df['Nota'].iloc[I] == result['Nota'].iloc[current_row] and taxas_df['IR_DT'].iloc[I] > 0:
                    row_data = [result['Corretora'].iloc[current_row],result['CPF'].iloc[current_row],result['Nota'].iloc[current_row],result['Data'].iloc[current_row],'V',result['Papel'].iloc[current_row],'DayTrade',0,0,0,0,0,taxas_df['IR_DT'].iloc[I]]
                    note_data.append(row_data)
                    break
    # ===================================================================================================
    # As operações de DayTrade com sobras deverão ter esse excesso adicionados nas operações Normais
    # Por exemplo: Compra 600 - Venda 400. 200 será inserido em uma operação Normal 
    # ===================================================================================================           
    for i in result.index:
        qtde = result['QTDE'].iloc[i]
        nota = result['Nota'].iloc[i]
        papel = result['Papel'].iloc[i]
        if qtde > 0:
            for I in grouped.index:
                if grouped['Nota'].iloc[I] == nota and grouped['Papel'].iloc[I] == papel and grouped['Operacao'].iloc[I] == 'DayTrade' and grouped['C/V'].iloc[I] == 'C':
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'C',grouped['Papel'].iloc[I],'DayTrade',0,qtde*(-1),(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),0,0]
                    note_data.append(row_data)
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'C',grouped['Papel'].iloc[I],'Normal',grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I],qtde,(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,0,0]
                    note_data.append(row_data)
                    print(f'Houve operação de Daytrade do papel {papel} com sobra de {qtde} açoes para Swing Trade, na operação de COMPRA.')
                    log += 'Houve operação de Daytrade do papel "'+papel+'" com sobra de "'+str(int(qtde))+'" açoes para Swing Trade, na operação de COMPRA.\n'
                    break
        elif qtde < 0:
            qtde = qtde*(-1)
            for I in grouped.index:
                if grouped['Nota'].iloc[I] == nota and grouped['Papel'].iloc[I] == papel and grouped['Operacao'].iloc[I] == 'DayTrade' and grouped['C/V'].iloc[I] == 'V':
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'V',grouped['Papel'].iloc[I],'DayTrade',0,qtde*(-1),(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),0,0]
                    note_data.append(row_data)
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'V',grouped['Papel'].iloc[I],'Normal',grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I],qtde,(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,0,((grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde)*0.00005]
                    note_data.append(row_data)
                    print(f'Houve operação de Daytrade do papel {papel} com sobra de {qtde} açoes para Swing Trade, na operação de VENDA.')
                    log += 'Houve operação de Daytrade do papel "'+papel+'" com sobra de "'+str(int(qtde))+'" açoes para Swing Trade, na operação de VENDA.\n'
                    break
    return note_data,taxas_df,log

# ===================================================================================================
# Acrescentando os custos operacionais (Corretagem, Imposto e Outros)
# Essa função foi comentada nos scripts por não ter uma real utilidade
# ===================================================================================================
def custos_operacionais(grouped,taxas_df):
    grouped['freq'] = grouped.groupby('Nota')['Nota'].transform('count')
    custos_data = []
    for current_row in taxas_df.index:
        for I in grouped.index:
            if grouped['Nota'].iloc[I] == taxas_df['Nota'].iloc[current_row]:
                row_data = [taxas_df['Custos_Op'].iloc[current_row]]
                custos_data.append(row_data)
    grouped['Custos_Op'] = pd.DataFrame(data=custos_data)
    grouped['Custos_Fin'] = grouped['Custos_Fin'] + (grouped['Custos_Op'] / grouped['freq'])
    return grouped

# ===================================================================================================
# Obtendo o valor correto do preço médio de cada operação
# ===================================================================================================
def preco_medio_correcao(grouped):
    preco_medio_data = []
    for current_row in grouped.index:
        if grouped['C/V'].iloc[current_row] == 'C':
            preco_medio = (grouped['Total'].iloc[current_row] + grouped['Custos_Fin'].iloc[current_row]) / grouped['Quantidade'].iloc[current_row]
        else:
            preco_medio = (grouped['Total'].iloc[current_row] - grouped['Custos_Fin'].iloc[current_row]) / grouped['Quantidade'].iloc[current_row]
        row_data = [preco_medio]
        preco_medio_data.append(row_data)
    cols = ['PM']
    preco_medio_df = pd.DataFrame(data=preco_medio_data, columns=cols)
    return preco_medio_df['PM']

# ===================================================================================================
# Agrupa as operações por tipo de trade (normal ou daytrade) com correção 
# de compra/venda a maior no DayTrade
# ===================================================================================================
def agrupar_operacoes_correcao(grouped,cols):
    groups = grouped.groupby(grouped.Operacao)
    try:
        normal_df = groups.get_group("Normal")
    except KeyError:
        normal_df = pd.DataFrame(columns=cols)
    try:
        daytrade_df = groups.get_group("DayTrade")
    except KeyError:
        daytrade_df = pd.DataFrame(columns=cols)
    return normal_df,daytrade_df

# ===================================================================================================
# Verifica se a Nota de Corretagem já foi processada anteriormente
# ===================================================================================================
def verifica_nota_corretagem(folder_path,nome,item):
    if exists(folder_path+'/'+nome):
        print_atencao()
        print("Todas as Notas de Corretagem contidadas no arquivo",basename(item),"já foram contabilizadas anteriormente.")
        print("Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo",nome,"na pasta",folder_path,'.')
        print("Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba DADOS do arquivo COIR.xslm também devem ser apagados.\n")
        log = 'ATENÇÃO:\n'
        log += ' - Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" já foram contabilizadas anteriormente.\n'
        log += ' - Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo "'+nome+'" na pasta "'+folder_path+'".\n'
        log += ' - Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba "DADOS" do arquivo "COIR.xslm" também devem ser apagados.\n'
        return log

# ===================================================================================================
# Cria o caminho completo de pasta/subpasta para salvar o resultado do processamento
# ===================================================================================================
def move_resultado(folder_path,cpf,nome,item,pagebmf):
    from os import makedirs
    from shutil import copytree, ignore_patterns
    if exists(folder_path+'/'+nome):
        print_atencao()
        print("Todas as Notas de Corretagem contidadas no arquivo",basename(item),"já foram contabilizadas anteriormente.")
        print("Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo",nome,"na pasta",folder_path,'.')
        print("Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba DADOS do arquivo COIR.xslm também devem ser apagados.\n")
        log = 'ATENÇÃO:\n'
        log += ' - Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" já foram contabilizadas anteriormente.\n'
        log += ' - Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo "'+nome+'" na pasta "'+folder_path+'".\n'
        log += ' - Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba "DADOS" do arquivo "COIR.xslm" também devem ser apagados.\n'
        pagebmf = 0
    elif not exists('./Resultado/'+cpf):
        source = './Apoio/'
        destination = './Resultado/'+cpf
        copytree(source, destination, ignore=ignore_patterns('*.pyc', 'tmp*','*.csv','Backup','_Temp'))
        makedirs(folder_path)
        #log = 'Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n'
        log = ''
    elif not exists(folder_path):
        makedirs(folder_path)
        #log = 'Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n'
        log = ''
    else:
        #log = 'Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n'
        log = ''
    return log,pagebmf

# ===================================================================================================
# Disponibiliza os dados coletados em um arquivo .xlsx separado por mês para as operações à vista
# ===================================================================================================
def arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df):
    with pd.ExcelWriter(folder_path+'/'+nome) as writer:
        note_df.to_excel(writer, sheet_name="Completo", index = False, merge_cells=True)
        normal_df.to_excel(writer, sheet_name="Normal", index = False, merge_cells=True)
        daytrade_df.to_excel(writer, sheet_name="DayTrade", index = False, merge_cells=True)
        taxas_df.to_excel(writer, sheet_name="Taxas", index = False, merge_cells=True)

# ===================================================================================================
# Disponibiliza os dados coletados em um arquivo .xlsx separado por mês para as operações BM&F
# ===================================================================================================
def arquivo_separado_bmf(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df):
    from openpyxl import load_workbook
    book = load_workbook(folder_path+'/'+nome)
    writer = pd.ExcelWriter(folder_path+'/'+nome, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    note_df.to_excel(writer, sheet_name="Completo", startrow=writer.sheets['Completo'].max_row, index = False,header= False)
    normal_df.to_excel(writer, sheet_name="Normal", startrow=writer.sheets['Normal'].max_row, index = False,header= False)
    daytrade_df.to_excel(writer, sheet_name="DayTrade", startrow=writer.sheets['DayTrade'].max_row, index = False,header= False)
    taxas_df.to_excel(writer, sheet_name="Taxas", startrow=writer.sheets['Taxas'].max_row, index = False,header= False)
    writer.save()
    #with pd.ExcelWriter(folder_path+'/'+nome) as writer:
    #    note_df.to_excel(writer, sheet_name="Completo", index = False, merge_cells=True)
    #    normal_df.to_excel(writer, sheet_name="Normal", index = False, merge_cells=True)
    #    daytrade_df.to_excel(writer, sheet_name="DayTrade", index = False, merge_cells=True)
    #    taxas_df.to_excel(writer, sheet_name="Taxas", index = False, merge_cells=True)
        
# ===================================================================================================
# Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
# ===================================================================================================
def arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df):
    from openpyxl import load_workbook
    import xlwings as xw
    log = ''
    try:
        book = load_workbook(current_path+cpf+"/Completo.xlsx")
        writer = pd.ExcelWriter(current_path+cpf+"/Completo.xlsx", engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        note_df.to_excel(writer, sheet_name="Completo", startrow=writer.sheets['Completo'].max_row, index = False,header= False)
        normal_df.to_excel(writer, sheet_name="Normal", startrow=writer.sheets['Normal'].max_row, index = False,header= False)
        daytrade_df.to_excel(writer, sheet_name="DayTrade", startrow=writer.sheets['DayTrade'].max_row, index = False,header= False)
        taxas_df.to_excel(writer, sheet_name="Taxas", startrow=writer.sheets['Taxas'].max_row, index = False,header= False)
        writer.save()
   
        # ===================================================================================================
        # Utilização do pacote xlwings para manipular arquivos excel com macro. 
        # O openpyxl se mostrou ineficaz para essa tarefa.
        # https://www.dataquest.io/blog/python-excel-xlwings-tutorial/
        # ===================================================================================================                   
        app = xw.App(visible=False)
        wb = xw.Book(current_path+cpf+"/COIR.xlsb")                 
        
        # ===================================================================================================
        # Desproteger as planilhas no arquivo COIR.xlsm que serão atualizadas durante a executção do script. 
        # Após a atualização essas planilhas serão protegidas novamente
        # https://docs.xlwings.org/en/stable/api.html#xlwings.Book.macro
        # ===================================================================================================
        desproteger = app.macro('Desproteger')
        desproteger()
        
        sheet = wb.sheets['Normais_Dados']
        last_row = sheet.range(1,1).end('down').row
        if 0 < last_row < 1048576:
            last_row = last_row + 1
            sheet.range("A{row}".format(row=last_row)).options(index=False,header=False).value = normal_df
        else:   
            sheet.range("A2").options(index=False,header=False).value = normal_df     
        sheet = wb.sheets['DayTrade_Dados']
        last_row2 = sheet.range(1,1).end('down').row
        if 0 < last_row2 < 1048576:
            last_row2 = last_row2 + 1
            sheet.range("A{row}".format(row=last_row2)).options(index=False,header=False).value = daytrade_df
        else:
            sheet.range("A2").options(index=False,header=False).value = daytrade_df       
        proteger = app.macro('Proteger')
        proteger()
        wb.save()
        wb.close()
        app.quit()        
    except FileNotFoundError:
        with pd.ExcelWriter(current_path+cpf+"/Completo.xlsx") as writer:
            note_df.to_excel(writer, sheet_name="Completo", index = False,header= True)
            normal_df.to_excel(writer, sheet_name="Normal", index = False,header= True)
            daytrade_df.to_excel(writer, sheet_name="DayTrade", index = False,header= True)
            taxas_df.to_excel(writer, sheet_name="Taxas", index = False,header= True)
    except PermissionError:
        print_erro()
        print('\n','O arquivo',current_path+cpf+"/Completo.xlsx",'está aberto, por favor feche-o e tente novamente')
        log += 'ERRO:\n' 
        log += ' - O arquivo'+current_path+cpf+"/Completo.xlsx"+' está aberto, por favor feche-o e tente novamente\n'
        '''Criar um controle para não mover os arquivos na rotina move_saida'''
# ===================================================================================================
# Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
# Os arquivos serão movidos de ./Entrada para a pasta ./Saida/CPF/Corretora/Ano
# ===================================================================================================
def move_saida(cpf,corretora,ano,item):
    from os import makedirs
    log = ''
    destino_path = './Saida/'
    destino_prefix = cpf+'/'+corretora+'/'+ano
    path_destino = join(destino_path, destino_prefix)   
    if exists(path_destino):
        print(f'Movendo o arquivo {basename(item)} para a pasta {path_destino}','\n')
        log = 'Movendo o arquivo "'+basename(item)+'" para a pasta '+path_destino+'.\n'
    if not exists('./Saida/'+cpf):
        makedirs(path_destino)
        print(f'Movendo o arquivo {basename(item)} para a pasta {path_destino}','\n')
        log = 'Movendo o arquivo "'+basename(item)+'" para a pasta '+path_destino+'.\n'
    elif not exists(path_destino):
        makedirs(path_destino)
        print(f'Movendo o arquivo {basename(item)} para a pasta {path_destino}','\n')
        log = 'Movendo o arquivo "'+basename(item)+'" para a pasta '+path_destino+'.\n'        
    try:
        shutil.move(item, join(path_destino, basename(item)))
        #print('Movendo o arquivo "{}" para a pasta "{}"'.format(basename(item), (path_destino)),'\n')
    except PermissionError:
        #print('\n')
        print_atencao()
        print(f'O arquivo {basename(item)} está aberto e não pode ser movido para a pasta {path_destino}','\n')
        print('Feche-o e mova-o manualmente','\n')
        log += 'ATENÇÃO:\n'
        log += ' - O arquivo "'+basename(item)+'" está aberto e não pode ser movido para a pasta "'+path_destino+'".\n'
        log += ' - Feche-o e mova-o manualmente.\n'
    return log

# ===================================================================================================
# Cria um arquivo de LOG para armazenar os dados do processamento
# ===================================================================================================
def log_processamento(current_path,cpf,log):
    nome_log = 'log_'+cpf+'.txt'
    ordem = list(dict.fromkeys(log))
    log_sem_repeticao = ''
    
    for n in range(0,len(ordem)):
            log_sem_repeticao += ordem[n]

    try:
        arquivo = open(current_path+cpf+'/'+nome_log,'a')
    except FileNotFoundError:
        arquivo = open(current_path+cpf+'/'+nome_log, 'w+')
    
    arquivo.write(log_sem_repeticao + '\n')
    arquivo.close()

# ===================================================================================================
# Cria um arquivo de LOG para armazenar os dados do processamento
# ===================================================================================================
#def log_processamento_agora(current_path,cpf,log):
#    nome_log = 'log_'+cpf+'.txt'
#    try:
#        arquivo = open(current_path+cpf+'/'+nome_log,'a')
#    except FileNotFoundError:
#        arquivo = open(current_path+cpf+'/'+nome_log, 'w+')
#    #arquivo.write(str((log + '\n')))
#    print(log)
#    arquivo.write(log + '\n')
#    #w.write(str(recorded._content))
#    arquivo.close()

# ===================================================================================================
# Processamento de notas de corretagens das corretoras do grupo XP (XP, Rico e Clear)
# ===================================================================================================
def xp_rico_clear(corretora,filename,item,log,page,pagebmf=0,control=0):      
    # ===================================================================================================
    # Coleta de dados por área de informação - Extraindo os dados das operações na B3 
    # ===================================================================================================         
    # 50.947,428.028,73.259,564.134   - Nota e data do pregão:
    # 143.172,424.894,160.278,560.256 - CPF  
    # 240.603,32.194,448.109,561.0    - Informações de compra e venda:

    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages=page, encoding="utf-8", area=((50.947,428.028,73.259,564.134),(143.172,424.894,160.278,560.256),(240.603,32.194,448.109,561.0)))

    df = pd.concat(data,axis=0,ignore_index=True)
    df['Preço / Ajuste'] = df['Preço / Ajuste'].apply(sanitiza_moeda).astype('float')
    df['Valor Operação / Ajuste'] = df['Valor Operação / Ajuste'].apply(sanitiza_moeda).astype('float')
    df['Quantidade'] = df['Quantidade'].apply(sanitiza_moeda).astype('float')
    df['Nr. nota'] = df['Nr. nota'].apply(sanitiza_moeda).astype('float')
    df['Especificação do título'] = sanitiza_especificacao_titulo(df['Especificação do título'])    
    df['Obs. (*)'] = sanitiza_observacao(df['Obs. (*)'])    
    if 'Unnamed: 0' in df.columns:
        df['Unnamed: 0'] = df['Unnamed: 0'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 1' in df.columns:
        df['Unnamed: 1'] = df['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 2' in df.columns:
        df['Unnamed: 2'] = df['Unnamed: 2'].apply(sanitiza_moeda).astype('float')
    
    # ===================================================================================================
    # Coleta de dados por área de informação - Extraindo os dados das taxas e impostos 
    # ===================================================================================================         
    # 50.947,428.028,73.259,564.134 - nota e data do pregão:
    # 450.341,32.576,639.253,544.276 -  Resumo dos negócios, Resumo financeiro e Custos operacionais:

    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='all', encoding="utf-8", area=((53.178,428.995,71.772,561.382),(450.341,32.576,639.253,544.276)))
    df_gastos = pd.concat(data,axis=0,ignore_index=True)
    df_gastos['Nr. nota'] = df_gastos['Nr. nota'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 0' in df_gastos.columns:
        df_gastos['Unnamed: 0'] = df_gastos['Unnamed: 0'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 0'].fillna(0, inplace=True)
    if 'Unnamed: 1' in df_gastos.columns:       
        #print(df_gastos['Unnamed: 1'].to_markdown())
        df_gastos['Unnamed: 1'] = df_gastos['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 1'].fillna(0, inplace=True)
    lista = list(df_gastos[df_gastos['Resumo dos Negócios'].str.contains("Valor das operações",na=False)].index)
    note_taxa = []
    
    #Obtem o número da conta na corretora
    conta = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='1', encoding="utf-8", area=(160.278,426.541,179.616,520.253))
    conta = pd.concat(conta,axis=0,ignore_index=True)
    conta = conta['Unnamed: 0'].iloc[0].strip().lstrip('0')
    #print(conta)
        
    #Verifica se a Nota de Corretagem já foi processada anteriormente
    cpf = str(df['C.P.F./C.N.P.J/C.V.M./C.O.B.'][1])
    nome = conta + '_' + df_gastos['Data pregão'][0][6:10] + '_' + df_gastos['Data pregão'][0][3:5] + '.xlsx'
    current_path = './Resultado/'
    folder_prefix = str(df['C.P.F./C.N.P.J/C.V.M./C.O.B.'][1] + '/' + corretora + '/' + df_gastos['Data pregão'][0][6:10])
    folder_path = join(current_path, folder_prefix)
    if exists(folder_path+'/'+nome):
        log.append(verifica_nota_corretagem(folder_path,nome,item))
        log_processamento(current_path,cpf,log)
        return()
    
    for current_row in lista:
        nota = df_gastos['Nr. nota'].iloc[current_row-8]
        data = datetime.strptime(df_gastos['Data pregão'].iloc[current_row-8], '%d/%m/%Y').date()
        total = df_gastos['Unnamed: 0'].iloc[current_row]
        vendas = df_gastos['Unnamed: 0'].iloc[current_row-6]
        liquidacao = df_gastos['Unnamed: 1'].iloc[current_row-5]
        registro = df_gastos['Unnamed: 1'].iloc[current_row-4]
        emolumentos = df_gastos['Unnamed: 1'].iloc[current_row+1]
        corretagem = df_gastos['Unnamed: 1'].iloc[current_row+5]
        imposto = df_gastos['Unnamed: 1'].iloc[current_row+8]
        irrf = df_gastos['Unnamed: 1'].iloc[current_row+9]
        outros = df_gastos['Unnamed: 1'].iloc[current_row+10]
        ir_daytrade = str(df_gastos['Resumo dos Negócios'].iloc[current_row+10])
        if ir_daytrade != "nan":
            ir_daytrade = ir_daytrade.split("Projeção R$ ")[1]
            outros = df_gastos['Unnamed: 1'].iloc[current_row+11]
        else:
            ir_daytrade = "0"
        ir_daytrade = float(ir_daytrade.replace('.','').replace(',','.'))
        basecalculo = str(df_gastos['Resumo Financeiro'].iloc[current_row+9])
        if basecalculo != "nan":
            basecalculo = basecalculo.split("base R$")[1]
        else:
            basecalculo = "0"
        basecalculo = float(basecalculo.replace('.','').replace(',','.'))       
        row_data = [nota,data,total,vendas,liquidacao,registro,emolumentos,corretagem,imposto,outros,emolumentos+liquidacao+registro,corretagem+imposto+outros,irrf,ir_daytrade,basecalculo]                                        
        note_taxa.append(row_data)
    cols = ['Nota','Data','Total','Vendas','Liquidação','Registro','Emolumentos','Corretagem','Imposto','Outros','Custos_Fin','Custos_Op','IRRF','IR_DT','BaseCalculo']
    taxas_df = pd.DataFrame(data=note_taxa, columns=cols)
    taxas_df = taxas_df.drop_duplicates(subset='Nota', keep='last', ignore_index=True)
    cont_notas = len(taxas_df['Nota'])
    if cont_notas > 1:
        log.append('Serão processadas ' + str(cont_notas) + ' notas de corretagens do mercado à vista.\n')
    else:
        log.append('Será processada ' + str(cont_notas) + ' nota de corretagem do mercado à vista.\n')

    #Incluir aqui a etapa para obter lista de linhas de cada operação
    operacoes = list(df[df['Negociação'].str.contains("1-BOVESPA",na=False)].index)
    note_data = []
    numero_nota = 0
    cpf = ''
    nome = ''
    ano = ''
    temp = ''
    for current_row in operacoes:
        cell_value = df['Nr. nota'].iloc[current_row-2]
        if cell_value > 0:
            numero_nota = df['Nr. nota'].iloc[current_row-2]
            data = df['Data pregão'].iloc[current_row-2]
            if ano == '':
                cpf = df['C.P.F./C.N.P.J/C.V.M./C.O.B.'].iloc[current_row-1]
                nome = conta + '_' + data[6:10] + '_' + data[3:5] + '.xlsx'
                ano = data[6:10]
            data = datetime.strptime(df['Data pregão'].iloc[current_row-2], '%d/%m/%Y').date()
            
        #Tipo de operação (Compra ou Venda)
        c_v = df['C/V'].iloc[current_row].strip()

        #Nome do ativo no pregão
        stock_title = df['Especificação do título'].iloc[current_row].strip()

        operacao = df['Obs. (*)'].iloc[current_row]
        if operacao == "D":
            operacao = "DayTrade"
        else:
            operacao = "Normal"
            
        #Quantidade operada de cada ativo por nota de corretagem
        quantidade = quantidade_operada(df['Quantidade'].iloc[current_row],df['Unnamed: 0'].iloc[current_row] if 'Unnamed: 0' in df.columns else 0,df['Unnamed: 1'].iloc[current_row] if 'Unnamed: 1' in df.columns else 0,df['Unnamed: 2'].iloc[current_row] if 'Unnamed: 2' in df.columns else 0)

        #Valor total de cada operação por nota de corretagem
        valor_total = valor_total_ativo(df['Valor Operação / Ajuste'].iloc[current_row],df['Unnamed: 2'].iloc[current_row] if 'Unnamed: 2' in df.columns else 0 ,df['Unnamed: 1'].iloc[current_row] if 'Unnamed: 1' in df.columns else 0)
        
        # Preço unitário da operação de cada ativo por nota de corretagem
        preco_unitario = valor_total / quantidade
            
        #Dividindo os custos e o IRRF por operação
        custo_financeiro,irrf_operacao = custos_por_operacao(taxas_df,numero_nota,c_v,valor_total,operacao)
    
        #Susbstitui o nome do papel no pregão pelo seu respectivo código na B3 no padrão "XXXX3"
        stock_title,log_nome_pregao = nome_pregao(acoes, stock_title, data)
        if log_nome_pregao != temp:
            temp = log_nome_pregao
            log.append(log_nome_pregao)
        
        #Calculando o preço médio de cada operação
        pm = preco_medio(c_v,valor_total,custo_financeiro,quantidade)
        
        row_data = [corretora, cpf, numero_nota, data, c_v, stock_title, operacao, preco_unitario, quantidade, valor_total, custo_financeiro, pm, irrf_operacao]
        note_data.append(row_data)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
    note_df = pd.DataFrame(data=note_data, columns=cols)
       
    #Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
    grouped = agrupar(note_df)
        
    # Seleção de papel isento de IR (IRRF e IRPF). Apenas uma operação (um papel) está sendo analisada por NC       
    note_data,log_isecao = isencao_imposto_renda(taxas_df,grouped,note_data)
    log.append(log_isecao)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
    note_df = pd.DataFrame(data=note_data, columns=cols)

    # Refaz o agrupamento para atulizar os dados de preço e quantidade com a correção de compra/venda a maior no DayTrade 
    grouped = agrupar(note_df)
    
    # Agrupa as operações por tipo de operação (Normal ou Daytrade)
    try:
        normal_df,daytrade_df,result = agrupar_operacoes(grouped,cols)
        # Insere o valor do IR para as operações de Daytrade"
        note_data,taxas_df,log_daytrade_ir = daytrade_ir(result,taxas_df,note_data,grouped)
        log.append(log_daytrade_ir)
        cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
        note_df = pd.DataFrame(data=note_data, columns=cols)
        # Refaz o agrupamento para atualizar os dados de preço e quantidade por cada ativo comprado/vendido 
        grouped = agrupar(note_df)
    except ValueError:
        agrupar_operacoes(grouped,cols)
        #normal_df,daytrade_df = agrupar_operacoes(grouped,cols)    
        #excluir esse retorno e testar!!!
        
    # Acrescentando os custos operacionais (Corretagem, Imposto e Outros)
    grouped = custos_operacionais(grouped,taxas_df)
    
    # Obtendo o valor correto do preço unitário de cada operação
    grouped['Preço'] = grouped['Total'] / grouped['Quantidade']

    # Obtendo o valor correto do preço médio de cada operação
    grouped['PM'] = preco_medio_correcao(grouped)

    # Inseri o número da conta na corretora
    grouped.insert(2,"Conta",int(conta),True)
    taxas_df.insert(0,"Conta",int(conta),True)

    # Agrupa as operações por tipo de trade com correção de compra/venda a maior no DayTrade
    normal_df,daytrade_df = agrupar_operacoes_correcao(grouped,cols)
    
    # Cria o caminho completo de pastas/subpasta para salvar o resultado do processamento
    current_path = './Resultado/'
    folder_prefix = cpf+'/'+corretora+'/'+ano
    folder_path = join(current_path, folder_prefix)
    log_move_resultado,pagebmf = move_resultado(folder_path,cpf,nome,item,pagebmf)
    log.append(log_move_resultado)

    # Disponibiliza os dados coletados em um arquivo .xlsx separado por mês
    arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)   

    # Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
    arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df)
    
    # Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
    if pagebmf != 0 and control != 0:
        xp_rico_clear_bmf(corretora,filename,item,log,pagebmf,control)
    #    log_move_resultado,pagebmf = move_resultado(folder_path,cpf,nome,item,pagebmf)
    #    log.append(log_move_resultado)
    #else:
    #    log_move_resultado,pagebmf = move_resultado(folder_path,cpf,nome,item,pagebmf)
    #    log.append(log_move_resultado)
    
    # Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
    log_move_saida = move_saida(cpf,corretora,ano,item)
    log.append(log_move_saida)
    
    # Cria um arquivo de LOG para armazenar os dados do processamento
    log.append('Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n')
    log_processamento(current_path,cpf,log)

# ===================================================================================================
# Processamento de notas de corretagens BM&F das corretoras do grupo XP (XP, Rico e Clear)
# ===================================================================================================    
def xp_rico_clear_bmf(corretora,filename,item,log,page,control):
    # ===================================================================================================
    # Coleta de dados por área de informação - Extraindo os dados das operações na B3 
    # ===================================================================================================           
    # 46.484,442.159,68.797,561.90    - Nota e data do pregão
    # 127.553,439.928,148.378,562.647 - CPF
    # 171.434,29.378,618.428,566.366  - Informações de compra e venda

    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages=page, encoding="utf-8", area=((46.484,442.159,68.797,561.90),(127.553,439.928,148.378,562.647),(171.434,29.378,618.428,566.366)))
    
    df = pd.concat(data,axis=0,ignore_index=True)
    #print(df.to_markdown())
    df['Nr. nota'] = df['Nr. nota'].apply(sanitiza_moeda).astype('float')
    df['C/V'] = sanitiza_especificacao_titulo(df['C/V'])
    #df['Mercadoria'] = sanitiza_especificacao_titulo(df['Mercadoria'])
    df['Mercadoria'] = df['Mercadoria'].str.replace(' ','',regex=False)
    #df['Vencimento'] = df['Vencimento'].str.replace('@','',regex=False)
    df['Quantidade'] = df['Quantidade'].apply(sanitiza_moeda).astype('float')
    df['Preço/Ajuste'] = df['Preço/Ajuste'].apply(sanitiza_moeda).astype('float')
    df['Tipo Negócio'] = sanitiza_especificacao_titulo(df['Tipo Negócio'])
    df['Vlr de Operação/Ajuste'] = df['Vlr de Operação/Ajuste'].apply(sanitiza_moeda).astype('float')
    df['D/C'] = sanitiza_especificacao_titulo(df['D/C'])
    df['Taxa Operacional'] = df['Taxa Operacional'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 0' in df.columns: #CPF do investidor
        df['Unnamed: 0'] = sanitiza_especificacao_titulo(df['Unnamed: 0'])
    if 'Unnamed: 1' in df.columns:
        df['Unnamed: 1'] = df['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 2' in df.columns:
        df['Unnamed: 2'] = df['Unnamed: 2'].apply(sanitiza_moeda).astype('float')
        
    # ===================================================================================================
    # Coleta de dados por área de informação - Extraindo os dados das taxas e impostos 
    # ===================================================================================================         
    # 46.484,442.159,68.797,561.90   - Nota e data do pregão
    # 619.916,29.378,709.166,566.366 - Resumo dos negócios, Resumo financeiro e Custos operacionais
    
    #data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages=page, encoding="utf-8", area=((46.484,442.159,68.797,561.90),(619.916,29.378,709.166,566.366)))   
    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages=page, encoding="utf-8", area=((46.484,442.159,68.797,561.90),(619.916,30.122,716.603,564.878)))
    df_gastos = pd.concat(data,axis=0,ignore_index=True)
    #print(df_gastos.to_markdown())
   
    df_gastos['Nr. nota'] = df_gastos['Nr. nota'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 0' in df_gastos.columns:
        df_gastos['Unnamed: 0'] = df_gastos['Unnamed: 0'].apply(sanitiza_nota_bmf)
        df_gastos['Unnamed: 0'] = df_gastos['Unnamed: 0'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 0'].fillna(0, inplace=True) 
    if 'Unnamed: 1' in df_gastos.columns:
        df_gastos['Unnamed: 1'].fillna(0, inplace=True)
    if 'Unnamed: 2' in df_gastos.columns:
       df_gastos['Unnamed: 2'] = df_gastos['Unnamed: 2'].apply(sanitiza_nota_bmf)
       df_gastos['Unnamed: 2'] = df_gastos['Unnamed: 2'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 2'].fillna(0, inplace=True)
    if 'Unnamed: 3' in df_gastos.columns:
       df_gastos['Unnamed: 3'] = df_gastos['Unnamed: 3'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 3'].fillna(0, inplace=True)
    if 'Unnamed: 4' in df_gastos.columns:
       df_gastos['Unnamed: 4'] = df_gastos['Unnamed: 4'].apply(sanitiza_nota_bmf)
       df_gastos['Unnamed: 4'] = df_gastos['Unnamed: 4'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 4'].fillna(0, inplace=True)
    if 'Unnamed: 5' in df_gastos.columns:
       df_gastos['Unnamed: 5'] = df_gastos['Unnamed: 5'].apply(sanitiza_nota_bmf)
       df_gastos['Unnamed: 5'] = df_gastos['Unnamed: 5'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 5'].fillna(0, inplace=True)
    if 'Unnamed: 6' in df_gastos.columns:
       df_gastos['Unnamed: 6'] = df_gastos['Unnamed: 6'].apply(sanitiza_nota_bmf)
       df_gastos['Unnamed: 6'] = df_gastos['Unnamed: 6'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 6'].fillna(0, inplace=True)
    if 'Unnamed: 7' in df_gastos.columns:
       df_gastos['Unnamed: 7'] = df_gastos['Unnamed: 7'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 7'].fillna(0, inplace=True)
    if 'Unnamed: 8' in df_gastos.columns:
       df_gastos['Unnamed: 8'] = df_gastos['Unnamed: 8'].apply(sanitiza_moeda).astype('float')
       df_gastos['Unnamed: 8'].fillna(0, inplace=True)
    lista = list(df_gastos[df_gastos['Venda disponível'].str.contains("IRRF",na=False)].index)
    note_taxa = []
    #print(df_gastos[['Unnamed: 0','Data pregão','Venda disponível', 'Unnamed: 1','Unnamed: 2','Unnamed: 3','Unnamed: 4','Venda Opções','Unnamed: 5']].head(40))
    #print(df_gastos[['Unnamed: 0', 'Unnamed: 1','Unnamed: 2','Unnamed: 3','Unnamed: 4','Venda Opções','Unnamed: 5']].tail(20))
    
    #Obtem o número da conta na corretora
    try:
        conta = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='1', encoding="utf-8", area=(160.278,426.541,179.616,520.253))
        conta = pd.concat(conta,axis=0,ignore_index=True)
        conta = conta['Unnamed: 0'].iloc[0].strip().lstrip('0')
    except KeyError:    
        conta = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='1', encoding="utf-8", area=(146.147,442.159,166.972,561.159))
        conta = pd.concat(conta,axis=0,ignore_index=True)
        conta = conta['Unnamed: 0'].iloc[0].strip().lstrip('0')
    
    #Verifica se a Nota de Corretagem já foi processada anteriormente
    if control == 2:   
        cpf = df['Unnamed: 0'].iloc[1]
        nome = conta + '_' + df_gastos['Data pregão'][0][6:10] + '_' + df_gastos['Data pregão'][0][3:5] + '.xlsx'
        current_path = './Resultado/'
        folder_prefix = (df['Unnamed: 0'].iloc[1] + '/' + corretora + '/' + df_gastos['Data pregão'][0][6:10])
        folder_path = join(current_path, folder_prefix)
        if exists(folder_path+'/'+nome):
            log.append(verifica_nota_corretagem(folder_path,nome,item))
            log_processamento(current_path,cpf,log)
            return()
    
    for current_row in lista: 
        nota = df_gastos['Nr. nota'].iloc[current_row-2]
        data = datetime.strptime(df_gastos['Data pregão'].iloc[current_row-2], '%d/%m/%Y').date()
        irrf = str(df_gastos['Unnamed: 1'].iloc[current_row+1])
        if irrf != "nan":
            irrf = irrf.split("|")[0]
            irrf = float(irrf.replace('.','').replace(',','.'))
        else:
            irrf = "0"       
        if irrf > 0:
            venda_disponivel = df_gastos['Unnamed: 1'].iloc[current_row-1]
            venda_disponivel = float(venda_disponivel.replace('.','').replace(',','.'))
            compra_disponivel = df_gastos['Unnamed: 3'].iloc[current_row-1]
            venda_opcoes = df_gastos['Unnamed: 4'].iloc[current_row-1]
            compra_opcoes = df_gastos['Unnamed: 5'].iloc[current_row-1]
            valor_negocios = df_gastos['Unnamed: 6'].iloc[current_row-1]
            ir_daytrade = df_gastos['Unnamed: 3'].iloc[current_row+1]
            corretagem = df_gastos['Unnamed: 4'].iloc[current_row+1]
            taxa_registro = df_gastos['Unnamed: 5'].iloc[current_row+1]
            emolumentos = df_gastos['Unnamed: 6'].iloc[current_row+1]
            outros_custos = df_gastos['Compra disponível'].iloc[current_row+3]
            outros_custos = float(outros_custos.replace('.','').replace(',','.'))
            imposto = df_gastos['Unnamed: 3'].iloc[current_row+3]
            ajuste_posicao = df_gastos['Unnamed: 4'].iloc[current_row+3]
            ajuste_daytrade = df_gastos['Unnamed: 5'].iloc[current_row+3]
            total_custos_operacionais = df_gastos['Unnamed: 6'].iloc[current_row+3]
            outros = df_gastos['Unnamed: 0'].iloc[current_row+5]
            ir_operacional = df_gastos['Unnamed: 1'].iloc[current_row+5]
            total_conta_investimento = df_gastos['Unnamed: 3'].iloc[current_row+5]#Fazer o mesmo procedimento feito no caso do IRRF
            total_conta_normal = df_gastos['Unnamed: 4'].iloc[current_row+5]
            total_liquido = df_gastos['Unnamed: 5'].iloc[current_row+5]
            total_liquido_nota = df_gastos['Unnamed: 6'].iloc[current_row+5]
        else:
            venda_disponivel = df_gastos['Unnamed: 2'].iloc[current_row-1]
            compra_disponivel = df_gastos['Unnamed: 4'].iloc[current_row-1]
            venda_opcoes = df_gastos['Unnamed: 5'].iloc[current_row-1]
            compra_opcoes = df_gastos['Unnamed: 6'].iloc[current_row-1]
            valor_negocios = df_gastos['Unnamed: 7'].iloc[current_row-1]
            ir_daytrade = df_gastos['Unnamed: 4'].iloc[current_row+1]
            corretagem = df_gastos['Unnamed: 5'].iloc[current_row+1]
            taxa_registro = df_gastos['Unnamed: 6'].iloc[current_row+1]
            emolumentos = df_gastos['Unnamed: 7'].iloc[current_row+1]
            outros_custos = df_gastos['Compra disponível'].iloc[current_row+3]
            outros_custos = float(outros_custos.replace('.','').replace(',','.'))
            imposto = df_gastos['Unnamed: 4'].iloc[current_row+3]
            ajuste_posicao = df_gastos['Unnamed: 5'].iloc[current_row+3]
            ajuste_daytrade = df_gastos['Unnamed: 6'].iloc[current_row+3]
            total_custos_operacionais = df_gastos['Unnamed: 7'].iloc[current_row+3]
            outros = df_gastos['Unnamed: 0'].iloc[current_row+5]
            ir_operacional = df_gastos['Unnamed: 2'].iloc[current_row+5]
            total_conta_investimento = df_gastos['Unnamed: 4'].iloc[current_row+5]#Fazer o mesmo procedimento feito no caso do IRRF
            total_conta_normal = df_gastos['Unnamed: 5'].iloc[current_row+5]
            total_liquido = df_gastos['Unnamed: 6'].iloc[current_row+5]
            total_liquido_nota = df_gastos['Unnamed: 7'].iloc[current_row+5]
        liquidacao = 0
        basecalculo = 0
        row_data = [nota,data,compra_disponivel,venda_disponivel,liquidacao,taxa_registro,emolumentos,corretagem,imposto,outros,emolumentos+liquidacao+taxa_registro+imposto+outros,corretagem+imposto+outros,irrf,ir_daytrade,basecalculo] 
        note_taxa.append(row_data)                                               
    cols = ['Nota','Data','Total','Vendas','Liquidação','Registro','Emolumentos','Corretagem','Imposto','Outros','Custos_Fin','Custos_Op','IRRF','IR_DT','BaseCalculo'] 
    taxas_df = pd.DataFrame(data=note_taxa, columns=cols)    
    indexNames = taxas_df[((taxas_df['Custos_Fin'] == 0) & (taxas_df['Custos_Op'] == 0))].index
    taxas_df.drop(indexNames ,inplace=True)
    taxas_df = taxas_df.drop_duplicates(subset=['Nota','Data'], keep='last', ignore_index=True)
    #taxas_df_remove = taxas_df.loc[((taxas_df['Custos_Fin'] == 0) & (taxas_df['Custos_Op'] == 0))]
    #taxas_df = taxas_df.drop(taxas_df_remove.index, inplace=True)   
    
    cont_notas = len(taxas_df['Nota'])
    if cont_notas > 1:
        log.append('Serão processadas ' + str(cont_notas) + ' notas de corretagens de Mercados Futuros ou  BMF.\n')
    else:
        log.append('Será processada ' + str(cont_notas) + ' nota de corretagem.\n')
    
    #Incluir aqui a etapa para obter lista de linhas de cada operação
    #operacoes = list(df[df['Tipo Negócio'].isin(['DAY TRADE','NORMAL','AJUPOS','TX. PERMANÊNCIA'])].index)#operacoes = list(df[df['Taxa Operacional'] > 0 ].index)
    operacoes = list(df[df['C/V'].isin(['C','V'])].index)#operacoes = list(df[df['Taxa Operacional'] > 0 ].index)
    vendas = list(df[df['C/V'].isin(['V'])].index)
        
    if len(operacoes) == 0 and control == 1:
        log.append('Nota(s) de Corretagem(ns) apenas com ajustes de posição, por isso não será contabilizada.\n')
        cpf = df['Unnamed: 0'].iloc[current_row-1]
        log_processamento(current_path,cpf,log)
        return
    elif len(operacoes) == 0 and control == 2:
        log.append('Nota(s) de Corretagem(ns) apenas com ajustes de posição, por isso não será contabilizada.\n')
        current_path = './Resultado/'
        cpf = df['Unnamed: 0'].iloc[current_row-1]
        data = df['Data pregão'].iloc[current_row-2]
        ano = data[6:10]
        nome = ''
        folder_prefix = cpf+'/'+corretora+'/'+ano
        folder_path = join(current_path, folder_prefix)      
        log_move_saida = move_saida(cpf,corretora,ano,item)
        log.append(log_move_saida)
        log_move_resultado,pagebmf = move_resultado(folder_path,cpf,nome,item,pagebmf=0)
        log.append(log_move_resultado) 
        log_processamento(current_path,cpf,log)
        return

    note_data = []
    numero_nota = 0
    cpf = ''
    nome = ''
    ano = ''
    temp = ''
    for current_row in operacoes:
        cell_value = df['Nr. nota'].iloc[current_row-2]               
        if cell_value > 0:
            numero_nota = df['Nr. nota'].iloc[current_row-2]
            data = df['Data pregão'].iloc[current_row-2]
            if ano == '':    
                cpf = df['Unnamed: 0'].iloc[current_row-1]
                nome = conta + '_' + data[6:10] + '_' + data[3:5] + '.xlsx'
                ano = data[6:10]
            data = datetime.strptime(df['Data pregão'].iloc[current_row-2], '%d/%m/%Y').date()
        
        if df['Tipo Negócio'].iloc[current_row] in 'NORMALDAY TRADE':
            #Tipo de operação (Compra ou Venda)        
            c_v = df['C/V'].iloc[current_row].strip()
                
            #Nome do ativo no pregão
            mercadoria = df['Mercadoria'].iloc[current_row].strip()
        
            tipo_negocio = df['Tipo Negócio'].iloc[current_row]
            operacao = df['Tipo Negócio'].iloc[current_row]
            if operacao == "DAY TRADE":
                operacao = "DayTrade"
            else:
                operacao = "Normal"
        
            #Preço unitário da operação de cada mercadoria por nota de corretagem
            preco_unitario = df['Preço/Ajuste'].iloc[current_row]
        
            #Quantidade operada de cada mercadoria por nota de corretagem
            quantidade = df['Quantidade'].iloc[current_row]
        
            #Valor total de cada operação por nota de corretagem
            valor_total,id,mult = mercadoria_ticket(mercadoria,preco_unitario,quantidade)
        
            #Valor de corretagem por cada mercadoria operada
            corretagem = df['Taxa Operacional'].iloc[current_row]
        
            #Alterao nome da variável para manter a compatibilidade com o script de ações
            stock_title = mercadoria
        
            
            #Dividindo os custos e o IRRF por operação
            custo_financeiro = 0
            if corretagem == 0:
                corretagem = 1
            for i in taxas_df.index:
                if taxas_df['Corretagem'].iloc[i] == 0:
                    taxas_df_corretagem = 1
                else:
                    taxas_df_corretagem = taxas_df['Corretagem'].iloc[i]
                if taxas_df['Custos_Fin'].iloc[i] == 0:
                    taxas_df_Custos_Fin = 1
                else:
                    taxas_df_Custos_Fin = taxas_df['Custos_Fin'].iloc[i]
                if numero_nota == taxas_df['Nota'].iloc[i]:
                    custo_financeiro = (corretagem/taxas_df_corretagem)*taxas_df_Custos_Fin
                    #custo_financeiro = (corretagem/taxas_df['Corretagem'].iloc[i])*taxas_df['Custos_Fin'].iloc[i]
                    break
            irrf_operacao = 0
            ir_daytrade = 0
        
            #Calculando o preço médio de cada operação - Para operações de Futuros não se caucula PM
            pm = 0
            
            row_data = [corretora, cpf, numero_nota, data, c_v, stock_title, operacao, preco_unitario, quantidade, valor_total, custo_financeiro + corretagem,pm,irrf_operacao,ir_daytrade,id,mult]
            note_data.append(row_data)
    
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço','Quantidade', 'Total','Custos_Fin','PM','IRRF','IR_DT','ID','FATOR']
    note_df = pd.DataFrame(data=note_data, columns=cols)
    
    #Contabiliza a quantidade de vendas nas operações DayTrade e Normal
    note_df = ir_bmf(cont_notas,note_df,taxas_df,row_data,note_data)
    
    #Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
    grouped = agrupar_bmf(note_df)
    #print(grouped.head(40))
    
    # Inseri o número da conta na corretora
    grouped.insert(2,"Conta",int(conta),True)
    taxas_df.insert(0,"Conta",int(conta),True)
    
    # Agrupa as operações por tipo de trade com correção de compra/venda a maior no DayTrade
    normal_df,daytrade_df = agrupar_operacoes_correcao(grouped,cols)

    # Cria o caminho completo de pastas/subpasta para salvar o resultado do processamento
    current_path = './Resultado/'
    folder_prefix = cpf+'/'+corretora+'/'+ano
    folder_path = join(current_path, folder_prefix)   
    if control == 2:
        log_move_resultado,pagebmf = move_resultado(folder_path,cpf,nome,item,pagebmf=1)
        log.append(log_move_resultado)

    # Disponibiliza os dados coletados em um arquivo .xlsx separado por mês
    if control == 2:
        arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)
    else:
        arquivo_separado_bmf(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)

    # Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
    arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df)
        
    # Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
    if control == 2:
        log_move_saida = move_saida(cpf,corretora,ano,item)
        log.append(log_move_saida)
    
    # Cria um arquivo de LOG para armazenar os dados do processamento
    if control == 2:
        log.append('Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n')
        log_processamento(current_path,cpf,log)
'''
    # Disponibiliza os dados coletados em um arquivo .xlsx separado por mês
    if control == 2 and pagebmf != 0:
        arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)
    elif control != 2:
        arquivo_separado_bmf(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)

    # Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
    if control == 2 and pagebmf != 0:
        arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df)
    elif control != 2:
        arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df)
        
    # Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
    if control == 2 and pagebmf != 0:
        log_move_saida = move_saida(cpf,corretora,ano,item)

    # Cria um arquivo de LOG para armazenar os dados do processamento
    if control == 2 and pagebmf != 0:
        log.append('Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n')
        log_processamento(current_path,cpf,log)
'''
# ===================================================================================================
# Processamento de notas de corretagens corretora "AGORA CORRETORA DE TITULOS E VALORES MOBILIARIOS S/A"
# ===================================================================================================
def agora(corretora,filename,item,log):
    #Extraindo os dados das operações na B3
    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='all', encoding="utf-8", area=((40.556,452.264,58.415,580.249),(126.128,453.008,145.474,578.761),(208.723,35.568,472.878,581.737)))
    df = pd.concat(data,axis=0,ignore_index=True)
    df['Preço / Ajuste'] = df['Preço / Ajuste'].apply(sanitiza_moeda).astype('float') 
    df['Valor Operação / Ajuste'] = df['Valor Operação / Ajuste'].apply(sanitiza_moeda).astype('float')
    df['Quantidade'] = df['Quantidade'].apply(sanitiza_moeda).astype('float')
    df['Nr.Nota'] = df['Nr.Nota'].apply(sanitiza_moeda).astype('float')
    df['Unnamed: 0'] = sanitiza_especificacao_titulo(df['Unnamed: 0'])
    df['Especificação do título'] = df['Especificação do título'] + df['Unnamed: 0']
    df['Obs. (*)'] = sanitiza_observacao(df['Obs. (*)'])
    if 'Unnamed: 1' in df.columns:
        df['Unnamed: 1'] = df['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 2' in df.columns:
        df['Unnamed: 2'] = df['Unnamed: 2'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 3' in df.columns:
        df['Unnamed: 3'] = df['Unnamed: 3'].apply(sanitiza_moeda).astype('float')
    
    #Extraindo os dados das taxas e impostos
    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='all', encoding="utf-8", area=((40.556,452.264,58.415,580.249),(472.131,35.568,667.086,566.855)))
    df_gastos = pd.concat(data,axis=0,ignore_index=True)
    df_gastos['Nr.Nota'] = df_gastos['Nr.Nota'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 0' in df_gastos.columns:
        df_gastos['Unnamed: 0'] = df_gastos['Unnamed: 0'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 0'].fillna(0, inplace=True)
    if 'Unnamed: 1' in df_gastos.columns:
        df_gastos['Unnamed: 1'] = df_gastos['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 1'].fillna(0, inplace=True)
    if 'Unnamed: 2' in df_gastos.columns:
        df_gastos['Unnamed: 2'] = df_gastos['Unnamed: 2'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 2'].fillna(0, inplace=True)
    lista = list(df_gastos[df_gastos['Resumo dos Negócios'].str.contains("Valor das operações",na=False)].index)
    note_taxa = []
    
    #Obtem o número da conta na corretora
    conta = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='1', encoding="utf-8", area=(144.727,455.017,164.818,539.845))
    conta = pd.concat(conta,axis=0,ignore_index=True)
    conta = conta['Unnamed: 0'].iloc[0].strip().lstrip('0')
    conta = conta.split(' -')[0] + conta.split(' -')[1]
    
    #Verifica se a Nota de Corretagem já foi processada anteriormente
    cpf = str(df['C.P.F./C.N.P.J./C.V.M./C.O.B.'][1])
    nome = conta + '_' + df_gastos['Data pregão'][0][6:10] + '_' + df_gastos['Data pregão'][0][3:5] + '.xlsx'
    current_path = './Resultado/'
    folder_prefix = str(df['C.P.F./C.N.P.J./C.V.M./C.O.B.'][1] + '/' +corretora + '/' + df_gastos['Data pregão'][0][6:10])
    folder_path = join(current_path, folder_prefix)
    if exists(folder_path+'/'+nome):
        log.append(verifica_nota_corretagem(folder_path,nome,item))
        log_processamento(current_path,cpf,log)
        return()
    
    for current_row in lista:
        nota = df_gastos['Nr.Nota'].iloc[current_row-8]
        data = datetime.strptime(df_gastos['Data pregão'].iloc[current_row-8], '%d/%m/%Y').date()
        total = df_gastos['Unnamed: 0'].iloc[current_row]
        vendas = df_gastos['Unnamed: 0'].iloc[current_row-6]
        if df_gastos['Unnamed: 2'].iloc[current_row-6] > 0:
            liquidacao = df_gastos['Unnamed: 2'].iloc[current_row-5]
            registro = df_gastos['Unnamed: 2'].iloc[current_row-4]
            emolumentos = df_gastos['Unnamed: 2'].iloc[current_row+1]
            corretagem = df_gastos['Unnamed: 2'].iloc[current_row+4] 
            imposto = df_gastos['Unnamed: 2'].iloc[current_row+6]
            irrf = df_gastos['Unnamed: 2'].iloc[current_row+7]
            basecalculo = str(df_gastos['Unnamed: 1'].iloc[current_row+7])
            if basecalculo != "nan":
                basecalculo = float(basecalculo.split("R$")[0])
            else:
                basecalculo = "0"
            outros = df_gastos['Unnamed: 2'].iloc[current_row+8]
            ir_daytrade = str(df_gastos['Resumo dos Negócios'].iloc[current_row+7])
            if ir_daytrade != "nan":
                ir_daytrade = ir_daytrade.split("Projeção R$ ")[1]
            else:
                ir_daytrade = "0"
            ir_daytrade = float(ir_daytrade.replace('.','').replace(',','.'))               
        else:
            liquidacao = df_gastos['Unnamed: 1'].iloc[current_row-5]
            registro = df_gastos['Unnamed: 1'].iloc[current_row-4]
            emolumentos = df_gastos['Unnamed: 1'].iloc[current_row+1]
            corretagem = df_gastos['Unnamed: 1'].iloc[current_row+4]
            imposto = df_gastos['Unnamed: 1'].iloc[current_row+6]
            irrf = 0
            basecalculo = 0
            outros = df_gastos['Unnamed: 1'].iloc[current_row+7]
            ir_daytrade = str(df_gastos['Resumo dos Negócios'].iloc[current_row+6])
            if ir_daytrade != "nan":
                ir_daytrade = ir_daytrade.split("Projeção R$ ")[1]
            else:
                ir_daytrade = "0"
            ir_daytrade = float(ir_daytrade.replace('.','').replace(',','.'))
        row_data = [nota,data,total,vendas,liquidacao,registro,emolumentos,corretagem,imposto,outros,emolumentos+liquidacao+registro,corretagem+imposto+outros,irrf,ir_daytrade,basecalculo]                                        
        note_taxa.append(row_data)
    cols = ['Nota','Data','Total','Vendas','Liquidação','Registro','Emolumentos','Corretagem','Imposto','Outros','Custos_Fin','Custos_Op','IRRF','IR_DT','BaseCalculo']
    taxas_df = pd.DataFrame(data=note_taxa, columns=cols)
    taxas_df = taxas_df.drop_duplicates(subset='Nota', keep='last', ignore_index=True)
     
    cont_notas = len(taxas_df['Nota'])
    if cont_notas > 1:
        log.append('Serão processadas ' + str(cont_notas) + ' notas de corretagens do mercado à vista.\n')
    else:
        log.append('Será processada ' + str(cont_notas) + ' nota de corretagem do mercado à vista.\n')
    
    #Incluir aqui a etapa para obter lista de linhas de cada operação
    operacoes = list(df[df['Negociação'].str.contains("BOVESPA",na=False)].index)
    note_data = []
    numero_nota = 0
    cpf = ''
    nome = ''
    ano = ''
    temp = ''
    for current_row in operacoes:
        cell_value = df['Nr.Nota'].iloc[current_row-2]
        if cell_value > 0:
            numero_nota = df['Nr.Nota'].iloc[current_row-2]
            data = df['Data pregão'].iloc[current_row-2]
            if ano == '':
                cpf = df['C.P.F./C.N.P.J./C.V.M./C.O.B.'].iloc[current_row-1]
                nome = conta + '_' + data[6:10] + '_' + data[3:5] + '.xlsx'
                ano = data[6:10]
            data = datetime.strptime(df['Data pregão'].iloc[current_row-2], '%d/%m/%Y').date()

        #Tipo de operação (Compra ou Venda)
        c_v = df['C/V'].iloc[current_row].strip()
    
        #Nome do ativo no pregão
        stock_title = df['Especificação do título'].iloc[current_row].strip()

        operacao = df['Obs. (*)'].iloc[current_row]
        if operacao == "D":
            operacao = "DayTrade"
        else:
            operacao = "Normal"
            #operacao = "DayTrade"
        
        #Quantidade operada de cada ativo por nota de corretagem
        quantidade = quantidade_operada(df['Quantidade'].iloc[current_row],df['Unnamed: 0'].iloc[current_row] if 'Unnamed: 0' in df.columns else 0,df['Unnamed: 1'].iloc[current_row] if 'Unnamed: 1' in df.columns else 0,df['Unnamed: 2'].iloc[current_row] if 'Unnamed: 2' in df.columns else 0)
        
        #Valor total de cada operação por nota de corretagem
        valor_total = valor_total_ativo(df['Valor Operação / Ajuste'].iloc[current_row],df['Unnamed: 2'].iloc[current_row] if 'Unnamed: 2' in df.columns else 0 ,df['Unnamed: 1'].iloc[current_row] if 'Unnamed: 1' in df.columns else 0)
 
        # Preço unitário da operação de cada ativo por nota de corretagem
        preco_unitario = valor_total / quantidade
    
        #Dividindo os custos e o IRRF por operação
        custo_financeiro,irrf_operacao = custos_por_operacao(taxas_df,numero_nota,c_v,valor_total,operacao)
        
        #Susbstitui o nome do papel no pregão pelo seu respectivo código na B3 no padrão "XXXX3"
        stock_title,log_nome_pregao = nome_pregao(acoes, stock_title, data)
        if log_nome_pregao != temp:
            temp = log_nome_pregao
            log.append(log_nome_pregao)
            
        #Calculando o preço médio de cada operação
        pm = preco_medio(c_v,valor_total,custo_financeiro,quantidade)
        
        row_data = [corretora, cpf, numero_nota, data, c_v, stock_title, operacao, preco_unitario, quantidade, valor_total, custo_financeiro, pm, irrf_operacao]
        note_data.append(row_data)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
    note_df = pd.DataFrame(data=note_data, columns=cols)
    
    #Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
    grouped = agrupar(note_df)
    
    # Seleção de papel isento de IR (IRRF e IRPF). Apenas uma operação (um papel) está sendo analisada por NC       
    note_data,log_isencao = isencao_imposto_renda(taxas_df,grouped,note_data)
    log.append(log_isencao)
    
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
    note_df = pd.DataFrame(data=note_data, columns=cols)
    
    # Refaz o agrupamento para atulizar os dados de preço e quantidade com a correção de compra/venda a maior no DayTrade
    grouped = agrupar(note_df)
        
    # Agrupa as operações por tipo de operação (Normal ou Daytrade)
    try:
        normal_df,daytrade_df,result = agrupar_operacoes(grouped,cols)
        # Insere o valor do IR para as operações de Daytrade"
        note_data,taxas_df,log_daytrade_ir = daytrade_ir(result,taxas_df,note_data,grouped)
        log.append(log_daytrade_ir)
        
        cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
        note_df = pd.DataFrame(data=note_data, columns=cols)
        # Refaz o agrupamento para atualizar os dados de preço e quantidade por cada ativo comprado/vendido
        grouped = agrupar(note_df)
    except ValueError:
        agrupar_operacoes(grouped,cols)
        #normal_df,daytrade_df = agrupar_operacoes(grouped,cols)
       
    # Acrescentando os custos operacionais (Corretagem, Imposto e Outros)
    grouped = custos_operacionais(grouped,taxas_df)
    
    # Obtendo o valor correto do preço unitário de cada operação 
    grouped['Preço'] = grouped['Total'] / grouped['Quantidade']
    
    # Obtendo o valor correto do preço médio de cada operação
    grouped['PM'] = preco_medio_correcao(grouped)
    
    # Inseri o número da conta na corretora
    grouped.insert(2,"Conta",int(conta),True)
    taxas_df.insert(0,"Conta",int(conta),True)
    
    # Agrupa as operações por tipo de trade com correção de compra/venda a maior no DayTrade 
    normal_df,daytrade_df = agrupar_operacoes_correcao(grouped,cols)

    # Cria o caminho completo de pastas/subpasta para salvar o resultado do processamento
    current_path = './Resultado/'
    folder_prefix = cpf+'/'+corretora+'/'+ano
    folder_path = join(current_path, folder_prefix)
    log_move_resultado,pagebmf = move_resultado(folder_path,cpf,nome,item,pagebmf=0)
    log.append(log_move_resultado)
    
    # Disponibiliza os dados coletados em um arquivo .xlsx separado por mês
    arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)
    
    # Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
    arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df)

    # Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
    log_move_saida = move_saida(cpf,corretora,ano,item)
    log.append(log_move_saida)

    # Cria um arquivo de LOG para armazenar os dados do processamento
    log.append('Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n')
    log_processamento(current_path,cpf,log)

# ===================================================================================================
# Processamento de notas de corretagens de uma corretora ainda não validada
# ===================================================================================================
def corretora_nao_validada(corretora,filename,item,log):
    #Extraindo os dados das operações na B3
    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='all', encoding="utf-8", area=((50.947,428.028,73.259,564.134),(143.172,424.894,160.278,560.256),(240.603,32.194,448.109,561.0)))
    df = pd.concat(data,axis=0,ignore_index=True)
    df['Preço / Ajuste'] = df['Preço / Ajuste'].apply(sanitiza_moeda).astype('float')
    df['Valor Operação / Ajuste'] = df['Valor Operação / Ajuste'].apply(sanitiza_moeda).astype('float')
    df['Quantidade'] = df['Quantidade'].apply(sanitiza_moeda).astype('float')
    df['Nr. nota'] = df['Nr. nota'].apply(sanitiza_moeda).astype('float')
    df['Especificação do título'] = sanitiza_especificacao_titulo(df['Especificação do título'])    
    df['Obs. (*)'] = sanitiza_observacao(df['Obs. (*)'])    
    if 'Unnamed: 0' in df.columns:
        df['Unnamed: 0'] = df['Unnamed: 0'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 1' in df.columns:
        df['Unnamed: 1'] = df['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 2' in df.columns:
        df['Unnamed: 2'] = df['Unnamed: 2'].apply(sanitiza_moeda).astype('float')

    #Extraindo os dados das taxas e impostos
    data = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='all', encoding="utf-8", area=((53.178,428.995,71.772,561.382),(450.341,32.576,639.253,544.276)))
    df_gastos = pd.concat(data,axis=0,ignore_index=True)     
    df_gastos['Nr. nota'] = df_gastos['Nr. nota'].apply(sanitiza_moeda).astype('float')
    if 'Unnamed: 0' in df_gastos.columns:
        df_gastos['Unnamed: 0'] = df_gastos['Unnamed: 0'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 0'].fillna(0, inplace=True)
    if 'Unnamed: 1' in df_gastos.columns:
        df_gastos['Unnamed: 1'] = df_gastos['Unnamed: 1'].apply(sanitiza_moeda).astype('float')
        df_gastos['Unnamed: 1'].fillna(0, inplace=True)
    lista = list(df_gastos[df_gastos['Resumo dos Negócios'].str.contains("Valor das operações",na=False)].index)
    note_taxa = []
   
    for current_row in lista:
        nota = df_gastos['Nr. nota'].iloc[current_row-8]
        data = datetime.strptime(df_gastos['Data pregão'].iloc[current_row-8], '%d/%m/%Y').date()
        total = df_gastos['Unnamed: 0'].iloc[current_row]
        vendas = df_gastos['Unnamed: 0'].iloc[current_row-6]
        liquidacao = df_gastos['Unnamed: 1'].iloc[current_row-5]
        registro = df_gastos['Unnamed: 1'].iloc[current_row-4]
        emolumentos = df_gastos['Unnamed: 1'].iloc[current_row+1]
        corretagem = df_gastos['Unnamed: 1'].iloc[current_row+5]
        imposto = df_gastos['Unnamed: 1'].iloc[current_row+8]
        irrf = df_gastos['Unnamed: 1'].iloc[current_row+9]
        outros = df_gastos['Unnamed: 1'].iloc[current_row+10]
        ir_daytrade = str(df_gastos['Resumo dos Negócios'].iloc[current_row+10])
        if ir_daytrade != "nan":
            ir_daytrade = ir_daytrade.split("Projeção R$ ")[1]
            outros = df_gastos['Unnamed: 1'].iloc[current_row+11]
        else:
            ir_daytrade = "0"
        ir_daytrade = float(ir_daytrade.replace('.','').replace(',','.'))
        basecalculo = str(df_gastos['Resumo Financeiro'].iloc[current_row+9])
        if basecalculo != "nan":
            basecalculo = basecalculo.split("base R$")[1]
        else:
            basecalculo = "0"
        basecalculo = float(basecalculo.replace('.','').replace(',','.'))       
        row_data = [nota,data,total,vendas,liquidacao,registro,emolumentos,corretagem,imposto,outros,emolumentos+liquidacao+registro,corretagem+imposto+outros,irrf,ir_daytrade,basecalculo]                                        
        note_taxa.append(row_data)
    cols = ['Nota','Data','Total','Vendas','Liquidação','Registro','Emolumentos','Corretagem','Imposto','Outros','Custos_Fin','Custos_Op','IRRF','IR_DT','BaseCalculo']
    taxas_df = pd.DataFrame(data=note_taxa, columns=cols)
    taxas_df = taxas_df.drop_duplicates(subset='Nota', keep='last', ignore_index=True)
    cont_notas = len(taxas_df['Nota'])
    if cont_notas > 1:
        log.append('Serão processadas ' + str(cont_notas) + ' notas de corretagens do mercado à vista.\n')
    else:
        log.append('Será processada ' + str(cont_notas) + ' nota de corretagem do mercado à vista.\n')

    #Incluir aqui a etapa para obter lista de linhas de cada operação
    operacoes = list(df[df['Negociação'].str.contains("1-BOVESPA",na=False)].index)
    note_data = []
    numero_nota = 0
    cpf = ''
    nome = ''
    ano = ''
    temp = ''
    
    for current_row in operacoes:
        cell_value = df['Nr. nota'].iloc[current_row-2]
        if cell_value > 0:
            numero_nota = df['Nr. nota'].iloc[current_row-2]
            data = df['Data pregão'].iloc[current_row-2]
            cpf = df['C.P.F./C.N.P.J/C.V.M./C.O.B.'].iloc[current_row-1]
            nome = data[6:10] + '_' + data[3:5] + '.xlsx'
            ano = data[6:10]
            data = datetime.strptime(df['Data pregão'].iloc[current_row-2], '%d/%m/%Y').date()

        #Tipo de operação (Compra ou Venda)
        c_v = df['C/V'].iloc[current_row].strip()

        #Nome do ativo no pregão
        stock_title = df['Especificação do título'].iloc[current_row].strip()

        operacao = df['Obs. (*)'].iloc[current_row]
        if operacao == "D":
            operacao = "DayTrade"
        else:
            operacao = "Normal"
        
        #Quantidade operada de cada ativo por nota de corretagem
        quantidade = quantidade_operada(df['Quantidade'].iloc[current_row],df['Unnamed: 0'].iloc[current_row] if 'Unnamed: 0' in df.columns else 0,df['Unnamed: 1'].iloc[current_row] if 'Unnamed: 1' in df.columns else 0,df['Unnamed: 2'].iloc[current_row] if 'Unnamed: 2' in df.columns else 0)

        #Valor total de cada operação por nota de corretagem
        valor_total = valor_total_ativo(df['Valor Operação / Ajuste'].iloc[current_row],df['Unnamed: 2'].iloc[current_row] if 'Unnamed: 2' in df.columns else 0 ,df['Unnamed: 1'].iloc[current_row] if 'Unnamed: 1' in df.columns else 0)
        
        # Preço unitário da operação de cada ativo por nota de corretagem
        preco_unitario = valor_total / quantidade
            
        #Dividindo os custos e o IRRF por operação
        custo_financeiro,irrf_operacao = custos_por_operacao(taxas_df,numero_nota,c_v,valor_total,operacao)
    
        #Susbstitui o nome do papel no pregão pelo seu respectivo código na B3 no padrão "XXXX3"
        stock_title,log_nome_pregao = nome_pregao(acoes, stock_title, data)
        if log_nome_pregao != temp:
            temp = log_nome_pregao
            log.append(log_nome_pregao)

        #Calculando o preço médio de cada operação
        pm = preco_medio(c_v,valor_total,custo_financeiro,quantidade)
        
        row_data = [corretora, cpf, numero_nota, data, c_v, stock_title, operacao, preco_unitario, quantidade, valor_total, custo_financeiro, pm, irrf_operacao]
        note_data.append(row_data)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
    note_df = pd.DataFrame(data=note_data, columns=cols)
    
    #Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
    grouped = agrupar(note_df)
        
    # Seleção de papel isento de IR (IRRF e IRPF). Apenas uma operação (um papel) está sendo analisada por NC       
    note_data,log_isecao = isencao_imposto_renda(taxas_df,grouped,note_data)
    log.append(log_isecao)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
    note_df = pd.DataFrame(data=note_data, columns=cols)

    # Refaz o agrupamento para atulizar os dados de preço e quantidade com a correção de compra/venda a maior no DayTrade 
    grouped = agrupar(note_df)
    
    # Agrupa as operações por tipo de operação (Normal ou Daytrade)
    try:
        normal_df,daytrade_df,result = agrupar_operacoes(grouped,cols)
        # Insere o valor do IR para as operações de Daytrade"
        note_data,taxas_df,log_daytrade_ir = daytrade_ir(result,taxas_df,note_data,grouped)
        log.append(log_daytrade_ir)
        cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço', 'Quantidade', 'Total', 'Custos_Fin', 'PM', 'IRRF']
        note_df = pd.DataFrame(data=note_data, columns=cols)
        # Refaz o agrupamento para atualizar os dados de preço e quantidade por cada ativo comprado/vendido 
        grouped = agrupar(note_df)
    except ValueError:
        agrupar_operacoes(grouped,cols)
        #normal_df,daytrade_df = agrupar_operacoes(grouped,cols)    

    # Acrescentando os custos operacionais (Corretagem, Imposto e Outros)
    '''grouped = custos_operacionais(grouped,taxas_df)'''
    
    # Obtendo o valor correto do preço unitário de cada operação
    grouped['Preço'] = grouped['Total'] / grouped['Quantidade']

    # Obtendo o valor correto do preço médio de cada operação
    grouped['PM'] = preco_medio_correcao(grouped)

    # Agrupa as operações por tipo de trade com correção de compra/venda a maior no DayTrade
    normal_df,daytrade_df = agrupar_operacoes_correcao(grouped,cols)
    
    # Cria o caminho completo de pastas/subpasta para salvar o resultado do processamento
    current_path = './Resultado/'
    folder_prefix = cpf+'/'+corretora+'/'+ano
    folder_path = join(current_path, folder_prefix)
    log_move_resultado = move_resultado(folder_path,cpf,nome,item,log,pagebmf=0)
    log.append(log_move_resultado)

    # Disponibiliza os dados coletados em um arquivo .xlsx separado por mês
    arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df)   

    # Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
    arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df)
        
    # Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
    log_move_saida = move_saida(cpf,corretora,ano,item)
    log.append(log_move_saida)
    
    # Cria um arquivo de LOG para armazenar os dados do processamento
    log.append('Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" foram processadas com sucesso.\n')
    log_processamento(current_path,cpf,log)
    
# ===================================================================================================
#                  Módulo principal - SISTEMA DE CONTROLE DE OPERAÇÕES E IRPF
#    Leitura, análise, extração, formatação e conversão das Notas de Corretagem no padrão SINANCOR
# ---------------------------------------------------------------------------------------------------
# Controle de arquivos que não foram processados durante a execução do script
# São arquivos que permanecerão na pasta ./Entrada após a execução do script
# ===================================================================================================
def extracao_nota_corretagem(path_origem='./Entrada/', ext='pdf'):
    from os import listdir
    resposta = ''
    for item in [join(path_origem, f) for f in listdir(path_origem) if isfile(join(path_origem, f)) and f.endswith(ext)]:
        filename = item
        log = []

        #Validação de notas de corretagem no padrão Sinacor
        try:        
            validacao = tabula.read_pdf(filename, pandas_options={'header': None}, guess=False, stream=True, multiple_tables=False, pages='1', silent=True, encoding="utf-8", area=(1.116,36.089,65.853,455.017))
            df_validacao = pd.concat(validacao,axis=1,ignore_index=True)
            df_validacao = pd.DataFrame({'NotaCorretagem': df_validacao[0].unique()})
            cell_value = df_validacao['NotaCorretagem'].iloc[0]
            if cell_value == 'NOTA DE NEGOCIAÇÃO' or cell_value == 'NOTA DE CORRETAGEM':
                print('processando o arquivo:',basename(item))
                log.append(datetime.today().strftime('%d/%m/%Y %H:%M:%S') + ' - Processando o arquivo "' + basename(item) + '"\n')
            else:
                print_atencao()
                print('O arquivo','"'+basename(item).upper()+'"','NÃO é uma Nota de Corretagem no Padrão Sinacor.','\n')
                continue
        except ValueError:
            print_atencao()
            print('O arquivo','"'+basename(item).upper()+'"','NÃO é uma Nota de Corretagem no Padrão Sinacor.','\n')
            continue

        #import pyPdf
        #from tabula import read_pdf
        #reader = pyPdf.PdfFileReader(open(filename, mode='rb' ))
        #n2 = reader.getNumPages() 
        #all_tables_stream = tabula.read_pdf(path, password = password, stream = "True", pages = n)

        #Validação de Corretoras cadastradas e implementadas, e corretoras não validadas
        corretora = tabula.read_pdf(filename, pandas_options={'dtype': str}, guess=False, stream=True, multiple_tables=True, pages='all', encoding="utf-8", area=(2.603,31.609,214.572,561.903)) 
        df_corretora = pd.concat(corretora,axis=0,ignore_index=True)
        corretora = tabula.read_pdf(filename, pages='1', **kwargs, area=(2.603,31.609,214.572,561.903))
        
        try:
            control,corretora,cell_value = valida_corretora(corretora)
            if control == 0:
                print('Corretora',cell_value, 'não implementada','\n')
                continue
            elif corretora in 'XPxpRICOricoCLEARclear' and control == 1:
                lista_acoes = list(df_corretora[df_corretora['NOTA DE NEGOCIAÇÃO'].str.contains(cell_value,na=False)].index)
                lista_bmf = list(df_corretora[df_corretora['Unnamed: 0'].str.contains(cell_value,na=False)].index)
                n1 = len(lista_acoes)
                n2 = len(lista_bmf)
                if n2 > 1:
                    page_acoes = ('1'+'-'+str(n1))
                    page_bmf = (str(int(n1+1))+'-'+str(int(n1+n2)))
                    #print(page_acoes,page_bmf)
                    xp_rico_clear(corretora,filename,item,log,page_acoes,page_bmf,control)
                else:
                    xp_rico_clear(corretora,filename,item,log,'all')
            elif corretora in 'XPxpRICOricoCLEARclear' and control == 2:
                xp_rico_clear_bmf(corretora,filename,item,log,'all',control=2)
            elif corretora in 'AGORAagora' and control == 1:
                agora(corretora,filename,item,log)
            elif control == 1:
                print(f'A corretora {corretora} ainda não foi validada.')
                print('Não há notas de corretagens suficientes para testá-la e implementá-la.')
                print('Todavia, será extraída com uma rotina de teste.')
                print('Dessa forma, Erros e inconsistência podem ocorrer durante o processamento.')
                if resposta == '':
                    while resposta not in 'SsNn':
                        resposta = str(input('Deseja realmente continuar [S/N]: '))
                if resposta in 'Ss':
                    corretora_nao_validada(corretora, filename,item,log)
                else:
                    continue                
            #elif control == 0:
            #    print('Corretora',cell_value, 'não implementada','\n')
            #    continue
        except ValueError as e:
            print(e)
            print('ValueError - Corretora ',cell_value, 'ocorreu erro durante o processamnto das notas de corretagens','\n')
            continue

# ===================================================================================================
# Mensagem de alerta para os aplicativos abertos do excel              
# O sistema continuará após a confirmação do uauário
# ===================================================================================================
def principal():
    print()
    print('-=' * 50)
    print(f'{"SISTEMA DE CONTROLE DE OPERAÇÕES E IRPF - COIR":^100}')
    print(f'{"Leitura, Análise, Extração, Formatação e Conversão das Notas de Corretagem no padrão SINANCOR":^100}')
    print('-=' * 50)
    print()
    excel_fechado = ' '
    print_atencao()
    print('Feche todos os documentos do Excel antes de iniciar o processamento das Notas de Corretagens.')
    print('Isso evitará erros e inconsistênica durante o processamento.\n')
    while excel_fechado not in 'SsNn':
        excel_fechado = str(input('O programa Excel está fechado [S/N]? ')).upper().strip()[0]
    if excel_fechado in 'Ss':
        print()
        print('-=' * 50)
        print('Iniciando o processamento das Notas de Corretagens...\n\n')
        extracao_nota_corretagem()

if __name__ == '__main__':
    principal()

print('-=' * 50)    
print('Fim do processamento!','\n')
input('Pressione qualquer tecla para concluir.\n')